import os
import random
import openpyxl
import xlsxwriter
import clickable_label
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox


class UiQuizMainWindow(object):
    def __init__(self):
        super().__init__()
        QuizMainWindow.setObjectName("QuizMainWindow")
        QuizMainWindow.resize(688, 690)
        QuizMainWindow.setGeometry(800, 300, 500, 500)
        self.quiz_main_widget = QtWidgets.QWidget(QuizMainWindow)
        self.quiz_main_widget.setObjectName("quiz_main_widget")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.quiz_main_widget)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.quiz_right_frame = QtWidgets.QFrame(self.quiz_main_widget)
        self.quiz_right_frame.setEnabled(True)
        self.quiz_right_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.quiz_right_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.quiz_right_frame.setObjectName("quiz_right_frame")
        self.gridLayout = QtWidgets.QGridLayout(self.quiz_right_frame)
        self.gridLayout.setObjectName("gridLayout")
        self.answer_a_button = QtWidgets.QPushButton(self.quiz_right_frame)
        self.answer_a_button.setStyleSheet("background-color: rgb(0, 0, 0);\n"
                                           "color: rgb(255, 255, 255);")
        self.answer_a_button.setObjectName("answer_a_button")
        self.gridLayout.addWidget(self.answer_a_button, 19, 0, 1, 2)
        self.label_a = QtWidgets.QLabel(self.quiz_right_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_a.setFont(font)
        self.label_a.setObjectName("label_a")
        self.gridLayout.addWidget(self.label_a, 17, 0, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem, 6, 0, 1, 1)
        self.main_question_text_display = QtWidgets.QTextBrowser(self.quiz_right_frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.main_question_text_display.sizePolicy().hasHeightForWidth())
        self.main_question_text_display.setSizePolicy(sizePolicy)
        self.main_question_text_display.setMinimumSize(QtCore.QSize(0, 50))
        self.main_question_text_display.setObjectName("main_question_text_display")
        self.gridLayout.addWidget(self.main_question_text_display, 7, 1, 1, 4)
        spacerItem1 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem1, 11, 0, 1, 1)
        self.line_2 = QtWidgets.QFrame(self.quiz_right_frame)
        self.line_2.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.gridLayout.addWidget(self.line_2, 18, 2, 1, 1)
        spacerItem2 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem2, 20, 0, 1, 1)
        self.label_possible_answers = QtWidgets.QLabel(self.quiz_right_frame)
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_possible_answers.setFont(font)
        self.label_possible_answers.setObjectName("label_possible_answers")
        self.gridLayout.addWidget(self.label_possible_answers, 16, 0, 1, 1)
        self.label_number_questions = QtWidgets.QLabel(self.quiz_right_frame)
        self.label_number_questions.setObjectName("label_number_questions")
        self.gridLayout.addWidget(self.label_number_questions, 3, 0, 1, 1)
        self.a_text_display = QtWidgets.QTextBrowser(self.quiz_right_frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(11)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.a_text_display.sizePolicy().hasHeightForWidth())
        self.a_text_display.setSizePolicy(sizePolicy)
        self.a_text_display.setMinimumSize(QtCore.QSize(220, 0))
        self.a_text_display.setObjectName("a_text_display")
        self.gridLayout.addWidget(self.a_text_display, 18, 0, 1, 2)
        self.label_quiz_details = QtWidgets.QLabel(self.quiz_right_frame)
        self.label_quiz_details.setObjectName("label_quiz_details")
        self.gridLayout.addWidget(self.label_quiz_details, 2, 0, 1, 5)
        self.label_d = QtWidgets.QLabel(self.quiz_right_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_d.setFont(font)
        self.label_d.setObjectName("label_d")
        self.gridLayout.addWidget(self.label_d, 24, 3, 1, 1)
        self.d_text_display = QtWidgets.QTextBrowser(self.quiz_right_frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(20)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.d_text_display.sizePolicy().hasHeightForWidth())
        self.d_text_display.setSizePolicy(sizePolicy)
        self.d_text_display.setMinimumSize(QtCore.QSize(220, 0))
        self.d_text_display.setObjectName("d_text_display")
        self.gridLayout.addWidget(self.d_text_display, 25, 3, 1, 2)
        self.answer_b_button = QtWidgets.QPushButton(self.quiz_right_frame)
        self.answer_b_button.setStyleSheet("background-color: rgb(0, 0, 0);\n"
                                           "color: rgb(255, 255, 255);")
        self.answer_b_button.setObjectName("Answer_b_button")
        self.gridLayout.addWidget(self.answer_b_button, 19, 3, 1, 2)
        spacerItem3 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem3, 23, 0, 1, 1)
        spacerItem4 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem4, 15, 0, 1, 1)
        self.answer_c_button = QtWidgets.QPushButton(self.quiz_right_frame)
        self.answer_c_button.setStyleSheet("background-color: rgb(0, 0, 0);\n"
                                           "color: rgb(255, 255, 255);")
        self.answer_c_button.setObjectName("answer_c_button")
        self.gridLayout.addWidget(self.answer_c_button, 26, 0, 1, 2)
        self.label_project_title_questionnaire = QtWidgets.QLabel(self.quiz_right_frame)
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.label_project_title_questionnaire.setFont(font)
        self.label_project_title_questionnaire.setObjectName("label_project_title_questionnaire")
        self.gridLayout.addWidget(self.label_project_title_questionnaire, 0, 0, 1, 1)
        spacerItem5 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem5, 22, 0, 1, 1)
        spacerItem6 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem6, 8, 0, 1, 1)
        self.c_text_display = QtWidgets.QTextBrowser(self.quiz_right_frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(10)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.c_text_display.sizePolicy().hasHeightForWidth())
        self.c_text_display.setSizePolicy(sizePolicy)
        self.c_text_display.setMinimumSize(QtCore.QSize(220, 0))
        self.c_text_display.setObjectName("c_text_display")
        self.gridLayout.addWidget(self.c_text_display, 25, 0, 1, 2)
        self.label_question = QtWidgets.QLabel(self.quiz_right_frame)
        self.label_question.setObjectName("label_question")
        self.gridLayout.addWidget(self.label_question, 7, 0, 1, 1)
        self.label_c = QtWidgets.QLabel(self.quiz_right_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_c.setFont(font)
        self.label_c.setObjectName("label_c")
        self.gridLayout.addWidget(self.label_c, 24, 0, 1, 1)
        self.line_3 = QtWidgets.QFrame(self.quiz_right_frame)
        self.line_3.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_3.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_3.setObjectName("line_3")
        self.gridLayout.addWidget(self.line_3, 25, 2, 1, 1)
        self.quiz_stats_table = QtWidgets.QTableWidget(self.quiz_right_frame)
        self.quiz_stats_table.setAlternatingRowColors(True)
        self.quiz_stats_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.quiz_stats_table.setGridStyle(QtCore.Qt.SolidLine)
        self.quiz_stats_table.setRowCount(1)
        self.quiz_stats_table.setObjectName("quiz_stats_table")
        self.quiz_stats_table.setColumnCount(4)
        item = QtWidgets.QTableWidgetItem()
        self.quiz_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.quiz_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.quiz_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.quiz_stats_table.setHorizontalHeaderItem(3, item)
        self.quiz_stats_table.horizontalHeader().setCascadingSectionResizes(True)
        self.quiz_stats_table.horizontalHeader().setDefaultSectionSize(110)
        self.quiz_stats_table.horizontalHeader().setSortIndicatorShown(False)
        self.quiz_stats_table.horizontalHeader().setStretchLastSection(True)
        self.quiz_stats_table.verticalHeader().setVisible(False)
        self.quiz_stats_table.verticalHeader().setDefaultSectionSize(35)
        self.quiz_stats_table.verticalHeader().setMinimumSectionSize(30)
        self.quiz_stats_table.verticalHeader().setSortIndicatorShown(False)
        self.quiz_stats_table.verticalHeader().setStretchLastSection(True)
        self.gridLayout.addWidget(self.quiz_stats_table, 29, 0, 1, 5)
        self.answer_d_button = QtWidgets.QPushButton(self.quiz_right_frame)
        self.answer_d_button.setStyleSheet("background-color: rgb(0, 0, 0);\n"
                                           "color: rgb(255, 255, 255);")
        self.answer_d_button.setObjectName("answer_d_button")
        self.gridLayout.addWidget(self.answer_d_button, 26, 3, 1, 2)
        spacerItem7 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem7, 13, 0, 1, 1)
        spacerItem8 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem8, 5, 0, 1, 1)
        self.label_stats = QtWidgets.QLabel(self.quiz_right_frame)
        self.label_stats.setObjectName("label_stats")
        self.gridLayout.addWidget(self.label_stats, 28, 0, 1, 1)
        self.label_b = QtWidgets.QLabel(self.quiz_right_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_b.setFont(font)
        self.label_b.setObjectName("label_b")
        self.gridLayout.addWidget(self.label_b, 17, 3, 1, 1)
        self.b_text_display = QtWidgets.QTextBrowser(self.quiz_right_frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(10)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.b_text_display.sizePolicy().hasHeightForWidth())
        self.b_text_display.setSizePolicy(sizePolicy)
        self.b_text_display.setMinimumSize(QtCore.QSize(220, 0))
        self.b_text_display.setObjectName("b_text_display")
        self.gridLayout.addWidget(self.b_text_display, 18, 3, 1, 2)
        self.label_average_difficulty = QtWidgets.QLabel(self.quiz_right_frame)
        self.label_average_difficulty.setObjectName("label_average_difficulty")
        self.gridLayout.addWidget(self.label_average_difficulty, 3, 3, 1, 1)
        spacerItem9 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem9, 21, 0, 1, 1)
        self.label_tags_question = QtWidgets.QLabel(self.quiz_right_frame)
        self.label_tags_question.setObjectName("label_tags_question")
        self.gridLayout.addWidget(self.label_tags_question, 9, 0, 1, 1)
        self.cancel_quiz_button = QtWidgets.QPushButton(self.quiz_right_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.cancel_quiz_button.setFont(font)
        self.cancel_quiz_button.setStyleSheet("background-color: rgb(255, 0, 0);")
        self.cancel_quiz_button.setObjectName("cancel_quiz_button")
        self.gridLayout.addWidget(self.cancel_quiz_button, 30, 4, 1, 1)
        self.finish_quiz_button = QtWidgets.QPushButton(self.quiz_right_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.finish_quiz_button.setFont(font)
        self.finish_quiz_button.setStyleSheet("background-color: rgb(0, 255, 0);")
        self.finish_quiz_button.setCheckable(False)
        self.finish_quiz_button.setAutoDefault(False)
        self.finish_quiz_button.setDefault(False)
        self.finish_quiz_button.setFlat(False)
        self.finish_quiz_button.setObjectName("finish_quiz_button")
        self.gridLayout.addWidget(self.finish_quiz_button, 30, 0, 1, 4)
        self.label_tag_1 = QtWidgets.QLabel(self.quiz_right_frame)
        self.label_tag_1.setEnabled(True)
        self.label_tag_1.setStyleSheet("background-color: rgb(0, 0, 0);\n"
                                       "color: rgb(255, 255, 255);")
        self.label_tag_1.setObjectName("label_tag_1")
        self.gridLayout.addWidget(self.label_tag_1, 9, 1, 1, 1)
        self.label_tag_2 = QtWidgets.QLabel(self.quiz_right_frame)
        self.label_tag_2.setStyleSheet("background-color: rgb(0, 0, 0);\n"
                                       "color: rgb(255, 255, 255);")
        self.label_tag_2.setObjectName("label_tag_2")
        self.gridLayout.addWidget(self.label_tag_2, 9, 3, 1, 1)
        self.gridLayout_2.addWidget(self.quiz_right_frame, 0, 2, 1, 1)
        self.quiz_left_frame = QtWidgets.QFrame(self.quiz_main_widget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(12)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.quiz_left_frame.sizePolicy().hasHeightForWidth())
        self.quiz_left_frame.setSizePolicy(sizePolicy)
        self.quiz_left_frame.setMinimumSize(QtCore.QSize(180, 0))
        self.quiz_left_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.quiz_left_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.quiz_left_frame.setObjectName("quiz_left_frame")
        self.formLayout = QtWidgets.QFormLayout(self.quiz_left_frame)
        self.formLayout.setObjectName("formLayout")
        self.new_quiz_button = QtWidgets.QPushButton(self.quiz_left_frame)
        self.new_quiz_button.setMinimumSize(QtCore.QSize(45, 50))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.new_quiz_button.setFont(font)
        self.new_quiz_button.setStyleSheet("background-color: rgb(255, 170, 0);")
        self.new_quiz_button.setObjectName("new_quiz_button")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.new_quiz_button)
        self.label_3 = QtWidgets.QLabel(self.quiz_left_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setWordWrap(True)
        self.label_3.setObjectName("label_3")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.label_3)
        self.restart_quiz_button = QtWidgets.QPushButton(self.quiz_left_frame)
        self.restart_quiz_button.setMinimumSize(QtCore.QSize(0, 50))
        self.restart_quiz_button.setStyleSheet("background-color: rgb(170, 255, 0);")
        self.restart_quiz_button.setObjectName("restart_quiz_button")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.restart_quiz_button)
        self.label_4 = QtWidgets.QLabel(self.quiz_left_frame)
        self.label_4.setWordWrap(True)
        self.label_4.setObjectName("label_4")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.label_4)
        self.label_project_nav = QtWidgets.QLabel(self.quiz_left_frame)
        self.label_project_nav.setWordWrap(True)
        self.label_project_nav.setObjectName("label_project_nav")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.SpanningRole, self.label_project_nav)

        spacerItem11 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.formLayout.setItem(6, QtWidgets.QFormLayout.LabelRole, spacerItem11)
        spacerItem12 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.formLayout.setItem(7, QtWidgets.QFormLayout.LabelRole, spacerItem12)
        spacerItem13 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.formLayout.setItem(8, QtWidgets.QFormLayout.LabelRole, spacerItem13)
        spacerItem14 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.formLayout.setItem(9, QtWidgets.QFormLayout.LabelRole, spacerItem14)
        self.main_list_widget = QtWidgets.QListWidget(self.quiz_left_frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.main_list_widget.sizePolicy().hasHeightForWidth())
        self.main_list_widget.setSizePolicy(sizePolicy)
        self.main_list_widget.setMinimumSize(QtCore.QSize(81, 0))
        self.main_list_widget.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.main_list_widget.setLineWidth(3)
        self.main_list_widget.setObjectName("main_list_widget")
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.SpanningRole, self.main_list_widget)
        self.select_current_quiz_button = QtWidgets.QPushButton(self.quiz_left_frame)
        self.select_current_quiz_button.setObjectName("change_quiz_button")
        self.formLayout.setWidget(4, QtWidgets.QFormLayout.SpanningRole, self.select_current_quiz_button)
        self.go_to_questionnaire_button = QtWidgets.QPushButton(self.quiz_left_frame)
        self.go_to_questionnaire_button.setObjectName("go_to_questionnaire_button")
        self.formLayout.setWidget(10, QtWidgets.QFormLayout.SpanningRole, self.go_to_questionnaire_button)
        self.label_3.raise_()
        self.new_quiz_button.raise_()
        self.restart_quiz_button.raise_()
        self.label_4.raise_()
        self.label_project_nav.raise_()
        self.main_list_widget.raise_()
        self.select_current_quiz_button.raise_()
        self.go_to_questionnaire_button.raise_()
        self.gridLayout_2.addWidget(self.quiz_left_frame, 0, 0, 1, 1)
        self.line = QtWidgets.QFrame(self.quiz_main_widget)
        self.line.setFrameShape(QtWidgets.QFrame.VLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.gridLayout_2.addWidget(self.line, 0, 1, 1, 1)
        QuizMainWindow.setCentralWidget(self.quiz_main_widget)
        self.quiz_menu_bar = QtWidgets.QMenuBar(QuizMainWindow)
        self.quiz_menu_bar.setGeometry(QtCore.QRect(0, 0, 688, 21))
        self.quiz_menu_bar.setObjectName("quiz_menu_bar")
        QuizMainWindow.setMenuBar(self.quiz_menu_bar)
        self.quiz_status_bar = QtWidgets.QStatusBar(QuizMainWindow)
        self.quiz_status_bar.setObjectName("quiz_status_bar")
        QuizMainWindow.setStatusBar(self.quiz_status_bar)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)

        self.textify(QuizMainWindow)
        QtCore.QMetaObject.connectSlotsByName(QuizMainWindow)

        # CLASS VARIABLES
        self.tried = False

        # UI ADDITIONS
        self.quiz_left_frame.setFixedWidth(180)
        self.label_tag_1.setAlignment(QtCore.Qt.AlignCenter)
        self.label_tag_2.setAlignment(QtCore.Qt.AlignCenter)
        self.delete_quiz_button = QtWidgets.QPushButton(self.quiz_left_frame)
        self.delete_quiz_button.setText("Delete quiz")
        self.formLayout.setWidget(5, QtWidgets.QFormLayout.SpanningRole, self.delete_quiz_button)
        self.delete_quiz_button.raise_()

        # CHILD WINDOW VARIABLE
        self.window = QtWidgets.QWidget()
        self.window.setGeometry(900, 400, 500, 500)

        # ADDITIONAL INIT
        for file in os.listdir():
            if file.startswith("Quiz"):
                self.main_list_widget.addItem(file[5:-5])
        if self.main_list_widget.count() == 0:
            self.quiz_right_frame.setEnabled(False)
        else:
            self.set_active_quiz(self.main_list_widget.count() - 1)

        # CONNECTIONS
        self.go_to_questionnaire_button.clicked.connect(self.switch_to_questionnaire)
        self.new_quiz_button.clicked.connect(self.new_quiz)
        self.select_current_quiz_button.clicked.connect(self.select_current_quiz)
        self.delete_quiz_button.clicked.connect(self.delete_quiz)
        self.answer_a_button.clicked.connect(lambda: self.check_correct_or_wrong("a"))
        self.answer_b_button.clicked.connect(lambda: self.check_correct_or_wrong("b"))
        self.answer_c_button.clicked.connect(lambda: self.check_correct_or_wrong("c"))
        self.answer_d_button.clicked.connect(lambda: self.check_correct_or_wrong("d"))

    def textify(self, QuizMainWindow):
        _translate = QtCore.QCoreApplication.translate
        QuizMainWindow.setWindowTitle(_translate("QuizMainWindow", "Quiz"))
        self.answer_a_button.setText(_translate("QuizMainWindow", "Answer a)"))
        self.label_a.setText(_translate("QuizMainWindow", "a)"))
        self.label_possible_answers.setText(_translate("QuizMainWindow", "Possible Answers"))
        self.label_number_questions.setText(_translate("QuizMainWindow", "# of questions:"))
        self.label_quiz_details.setText(_translate("QuizMainWindow", "Quiz Details:"))
        self.label_d.setText(_translate("QuizMainWindow", "d)"))
        self.answer_b_button.setText(_translate("QuizMainWindow", "Answer b)"))
        self.answer_c_button.setText(_translate("QuizMainWindow", "Answer c)"))
        self.label_project_title_questionnaire.setText(_translate("QuizMainWindow", "Quiz Title"))
        self.label_question.setText(_translate("QuizMainWindow", "Question:"))
        self.label_c.setText(_translate("QuizMainWindow", "c)"))
        item = self.quiz_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("QuizMainWindow", "Questions"))
        item = self.quiz_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("QuizMainWindow", "Answered"))
        item = self.quiz_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("QuizMainWindow", "Percentage"))
        item = self.quiz_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("QuizMainWindow", "Percentage Correct"))
        self.answer_d_button.setText(_translate("QuizMainWindow", "Answer d)"))
        self.label_stats.setText(_translate("QuizMainWindow", "Stats"))
        self.label_b.setText(_translate("QuizMainWindow", "b)"))
        self.label_average_difficulty.setText(_translate("QuizMainWindow", "Average Difficulty:"))
        self.label_tags_question.setText(_translate("QuizMainWindow", "Tags"))
        self.cancel_quiz_button.setText(_translate("QuizMainWindow", "(X) Cancel"))
        self.finish_quiz_button.setText(_translate("QuizMainWindow", "(✓) Finish Quiz"))
        self.label_tag_1.setText(_translate("QuizMainWindow", "Tag 1"))
        self.label_tag_2.setText(_translate("QuizMainWindow", "Tag 2"))
        self.new_quiz_button.setText(_translate("QuizMainWindow", "(+)"))
        self.label_3.setText(_translate("QuizMainWindow", "New Quiz \n"
                                                          "(This goes first!)"))
        self.restart_quiz_button.setText(_translate("QuizMainWindow", "(<--)"))
        self.label_4.setText(_translate("QuizMainWindow", "Restart current Quiz"))
        self.label_project_nav.setText(_translate("QuizMainWindow", "Click here to navigate through projects:"))
        self.select_current_quiz_button.setText(_translate("QuizMainWindow", "Select current quiz"))
        self.go_to_questionnaire_button.setText(_translate("QuizMainWindow", "Lets prepare more\n"
                                                                             " questionnaires!"))

    def new_quiz(self):
        if self.label_project_title_questionnaire.text() != "Questionnaire Title":
            # INSTANTIATION
            quiz_pop_up = CustomQuizPopUp()

    def set_active_quiz(self, index=0):
        # FETCHING DATA
        wb = openpyxl.load_workbook(f"Quiz_{self.main_list_widget.item(index).text()}.xlsx")
        sheet = wb.active

        # -- Fetching questions column in excel file, adding it to a list
        self.question_list = [cell.value for cell in sheet["A"] if cell.value != "QUESTIONS" and cell.value is not None]
        self.question_tag_list = [cell.value for cell in sheet["C"] if cell.value != "QUESTION TAGS"]
        self.answer_list = [cell.value for cell in sheet["D"] if cell.value != "ANSWER"]

        # -- Shuffling
        questions_and_tags = list(zip(self.question_list, self.question_tag_list, self.answer_list))
        random.shuffle(questions_and_tags)
        self.question_list, self.question_tag_list, self.answer_list = zip(*questions_and_tags)
        answer_list = list(self.answer_list)
        shuffle_list = [i for i in range(len(self.question_list))]
        random.shuffle(shuffle_list)
        current_question = "1"

        # UPDATE UI -- Updating labels and status bar to reflect fetched data
        self.label_number_questions.setText(f"""# of Questions: {len(self.question_list)} """)
        self.label_question.setText(f"""Question {current_question}/{len(self.question_list)}:""")
        self.quiz_status_bar.showMessage(f"Active questionnaire: "
                                         f"{self.main_list_widget.item(index).text()}")
        self.label_project_title_questionnaire.setText(self.main_list_widget.item(index).text())

        # -- Fetching difficulty values
        if sheet["H2"].value != "" and sheet["H2"].value is not None:  # DIFFICULTY
            diff_average = 0
            count = 0
            for cell in sheet["H"]:
                if cell.value != "DIFFICULTY" and cell.value is not None:
                    diff_average = diff_average + cell.value
                    count += 1
            average = round(diff_average / count, 2)
            self.label_average_difficulty.setText(f"Average Difficulty: {str(average)}")

        # -- Show question text and tags
        current_question = int(self.label_question.text().split("/")[0][9:])
        self.main_question_text_display.setPlainText(self.question_list[0])
        try:
            self.label_tag_1.setText(self.question_tag_list[0].split("//")[0])
        except IndexError:
            pass
        try:
            self.label_tag_2.setText(self.question_tag_list[0].split("//")[1])
        except IndexError:
            pass

        # -- Update stats table
        self.quiz_stats_table.setItem(0, 0, QtWidgets.QTableWidgetItem(str(len(self.question_list))))
        self.quiz_stats_table.setItem(0, 1, QtWidgets.QTableWidgetItem(str(
            int(self.label_question.text().split("/")[0][9:]) - 1)))
        questions = self.quiz_stats_table.item(0, 0).text()
        answered = self.quiz_stats_table.item(0, 1).text()
        if answered != "0":
            self.quiz_stats_table.setItem(0, 2, QtWidgets.QTableWidgetItem(
                str(round(int(questions) / int(answered), 2)) + "%"))
        else:
            self.quiz_stats_table.setItem(0, 2, QtWidgets.QTableWidgetItem("0%"))
        self.quiz_stats_table.setItem(0, 3, QtWidgets.QTableWidgetItem("-"))

        wb.close()

        # METHOD SERIALIZATION
        self.start_quiz_logic(answer_list)

    def select_current_quiz(self):
        # CHECK UI
        if len(self.main_list_widget.selectedItems()) > 0:
            item_list = [self.main_list_widget.item(i).text() for i in range(len(self.main_list_widget))]
            # UPDATE UI
            self.set_active_quiz(item_list.index(self.main_list_widget.selectedItems()[0].text()))
        else:
            self.generic_error("No quiz selected",
                               "Please select a quiz to switch to from the list shown above.")

    def start_quiz_logic(self, answer_list):
        # START LOGIC
        wb = openpyxl.load_workbook(f"Quiz_{self.label_project_title_questionnaire.text()}.xlsx")
        sheet = wb.active
        question = self.main_question_text_display.toPlainText()
        for cell in sheet["A"]:
            if cell.value == question:
                correct_answer = sheet["D" + cell.coordinate[1:]].value
                break
        wb.close()

        # -- Disable repeated answers
        index_list = [i for i in range(len(answer_list))]
        specific_index_list = []
        v = answer_list.index(correct_answer)
        specific_index_list.append(v)
        index_list.remove(v)

        x = random.choice(index_list)
        specific_index_list.append(x)
        index_list.remove(x)
        y = random.choice(index_list)
        specific_index_list.append(y)
        index_list.remove(y)
        z = random.choice(index_list)
        specific_index_list.append(z)
        index_list.remove(z)

        quiz_wrong_list = [answer_list[x], answer_list[y],
                           answer_list[z], correct_answer]
        a = random.choice(quiz_wrong_list)
        quiz_wrong_list.remove(a)
        b = random.choice(quiz_wrong_list)
        quiz_wrong_list.remove(b)
        c = random.choice(quiz_wrong_list)
        quiz_wrong_list.remove(c)
        d = random.choice(quiz_wrong_list)
        quiz_wrong_list.remove(d)

        self.a_text_display.setPlainText(str(a))
        self.b_text_display.setPlainText(str(b))
        self.c_text_display.setPlainText(str(c))
        self.d_text_display.setPlainText(str(d))

        common_style = "background-color: rgb(255, 255, 255);"
        letters = ["a", "b", "c", "d"]
        for i in letters:
            getattr(self, i + "_text_display").setStyleSheet(common_style)
        for i in ["a", "b", "c", "d"]:
            getattr(self, "answer_" + i + "_button").setEnabled(True)

    def delete_quiz(self):
        # SHOW CONFIRMATION WIDGET
        confirmation = QtWidgets.QMessageBox.question(self.window, "Confirmation",
                                                      f"Are you sure you want to permanently delete this quiz?:\n\n"
                                                      f"{self.main_list_widget.selectedItems()[0].text()}")
        # DELETE EXCEL FILE
        if confirmation == QMessageBox.Yes:
            try:
                deleted_quiz = f"Quiz_{self.main_list_widget.selectedItems()[0].text()}.xlsx"
                os.remove(deleted_quiz)
                quiz_ui.__init__()
            except IndexError:
                self.generic_error("No questionnaire selected", "Please select a questionnaire from the list above to "
                                                                "delete it")
        else:
            self.generic_information("", "Questionnaire was not deleted.")

    def check_correct_or_wrong(self, letter):
        # FETCH DATA
        wb = openpyxl.load_workbook(f"Quiz_{self.label_project_title_questionnaire.text()}.xlsx")
        sheet = wb.active
        question = self.main_question_text_display.toPlainText()
        # -- Get user answer
        user_answer = getattr(self, letter + "_text_display").toPlainText()
        # -- Get correct answer
        for cell in sheet["A"]:
            if cell.value == question:
                correct_answer = sheet["D" + cell.coordinate[1:]].value
                break
        questions = self.quiz_stats_table.item(0, 0).text()
        if not self.tried:
            tries = int(questions)
        else:
            tries = int(round(int(self.quiz_stats_table.item(0, 3).text().replace("%", "").split(".")[0]) / 100
                        * int(questions)))
        percentage_correct = (tries / int(questions)) * 100
        if user_answer == correct_answer:
            # CORRECT ANSWER LOGIC
            current_question = int(self.label_question.text().split("/")[0][9:]) + 1
            self.label_question.setText(f"""Question {current_question}/{self.quiz_stats_table.item(0, 0).text()}:""")
            # -- Show question text and tags
            current_question = int(self.label_question.text().split("/")[0][9:]) - 1

            # -- Stop when we've reached last question
            if current_question == int(self.label_number_questions.text()[16:]):
                self.generic_information("Quiz Completed", "Quiz was completed with the following score:\n")
                return None

            # -- Display current question if we haven't reached last question
            self.main_question_text_display.setPlainText(self.question_list[current_question])
            try:
                self.label_tag_1.setText(self.question_tag_list[current_question].split("//")[0])
            except IndexError:
                pass
            try:
                self.label_tag_2.setText(self.question_tag_list[current_question].split("//")[1])
            except IndexError:
                pass

            # -- Update Stats table
            self.quiz_stats_table.setItem(0, 1, QtWidgets.QTableWidgetItem(str(
                int(self.label_question.text().split("/")[0][9:]) - 1)))

            answered = self.quiz_stats_table.item(0, 1).text()
            if answered != "0":
                self.quiz_stats_table.setItem(0, 2, QtWidgets.QTableWidgetItem(
                    str(round((int(answered) / int(questions) * 100), 2)) + "%"))
                self.quiz_stats_table.setItem(0, 3, QtWidgets.QTableWidgetItem(
                    str(round(percentage_correct, 2)) + "%"))

            # -- Color correct
            correct_style_sheet = "background-color: rgb(50, 255, 50);"
            getattr(self, letter + "_text_display").setStyleSheet(correct_style_sheet)

        # -- Disable all answer buttons
            for i in ["a", "b", "c", "d"]:
                getattr(self, "answer_" + i + "_button").setEnabled(False)

            QtCore.QTimer.singleShot(2000, lambda: self.start_quiz_logic(self.answer_list))

        else:
            # -- Update Stats table # TODO: Fix numbering, somehow its sometimes advancing by two
            current_question = int(self.label_question.text().split("/")[0][9:]) + 1
            self.label_question.setText(f"""Question {current_question}/{self.quiz_stats_table.item(0, 0).text()}:""")
            # -- Show question text and tags
            current_question = int(self.label_question.text().split("/")[0][9:]) - 1

            # -- Stop when we've reached last question
            if current_question == int(self.label_number_questions.text()[16:]):
                self.generic_information("Quiz Completed", "Quiz was completed with the following score:\n")
                return None

            self.quiz_stats_table.setItem(0, 1, QtWidgets.QTableWidgetItem(str(
                int(self.label_question.text().split("/")[0][9:]))))
            answered = self.quiz_stats_table.item(0, 1).text()
            if answered != "0":
                self.quiz_stats_table.setItem(0, 2, QtWidgets.QTableWidgetItem(
                    str(round((int(answered) / int(questions) * 100), 2)) + "%"))
            tries -= 1
            self.tried = True
            percentage_correct = (tries / int(questions)) * 100
            self.quiz_stats_table.setItem(0, 3, QtWidgets.QTableWidgetItem(
                str(round(percentage_correct, 2)) + "%"))

            # -- Color wrong
            wrong_style_sheet = "background-color: rgb(255, 50, 50);"
            getattr(self, letter + "_text_display").setStyleSheet(wrong_style_sheet)

            for i in ["a", "b", "c", "d"]:
                if getattr(self, i + "_text_display").toPlainText() == correct_answer:
                    correct_letter = i
                    break
            correct_style_sheet = "background-color: rgb(50, 255, 50);"
            getattr(self, correct_letter + "_text_display").setStyleSheet(correct_style_sheet)

            # -- Disable all answer buttons
            for i in ["a", "b", "c", "d"]:
                getattr(self, "answer_" + i + "_button").setEnabled(False)

            QtCore.QTimer.singleShot(2000, lambda: self.start_quiz_logic(self.answer_list))

        wb.close()

    @staticmethod
    def switch_to_questionnaire():
        QuizMainWindow.hide()
        questionnaireMainWindow.show()

    @staticmethod
    def generic_information(title="Information", text="Details", icon=QMessageBox.Information):
        parent = QtWidgets.QMessageBox()
        parent.setGeometry(800, 300, 500, 500)
        parent.setWindowTitle(title)
        parent.setText(text)
        parent.setIcon(icon)
        parent.exec_()

    @staticmethod
    def generic_error(title="Error", text="Error Details", icon=QMessageBox.Warning):
        parent = QtWidgets.QMessageBox()
        parent.setGeometry(800, 300, 500, 500)
        parent.setWindowTitle(title)
        parent.setText(text)
        parent.setIcon(icon)
        parent.exec_()


class CustomQuizPopUp(QtWidgets.QDialog):
    def __init__(self):
        super().__init__()
        CustomQuizPopUp.setWindowTitle(self, "Quiz Options")

        self.setGeometry(250, 500, 700, 300)
        self.resize(324, 309)
        self.gridLayout = QtWidgets.QGridLayout(self)
        self.frame = QtWidgets.QFrame(self)
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.gridLayout_2 = QtWidgets.QGridLayout(self.frame)
        self.question_difficulty_hSlider = QtWidgets.QSlider(self.frame)
        self.question_difficulty_hSlider.setOrientation(QtCore.Qt.Horizontal)
        self.gridLayout_2.addWidget(self.question_difficulty_hSlider, 7, 1, 1, 1)
        self.exclude_tag_button = QtWidgets.QPushButton(self.frame)
        self.exclude_tag_button.setEnabled(False)
        self.gridLayout_2.addWidget(self.exclude_tag_button, 13, 0, 1, 1)
        self.label_percentage_max_questions = QtWidgets.QLabel(self.frame)
        self.gridLayout_2.addWidget(self.label_percentage_max_questions, 4, 0, 1, 1)
        self.label_quiz_summary = QtWidgets.QLabel(self.frame)
        self.gridLayout_2.addWidget(self.label_quiz_summary, 0, 0, 1, 2)
        self.label_excluded_tags = QtWidgets.QLabel(self.frame)
        self.gridLayout_2.addWidget(self.label_excluded_tags, 12, 1, 1, 1)
        self.question_difficulty_comboBox = QtWidgets.QComboBox(self.frame)
        self.gridLayout_2.addWidget(self.question_difficulty_comboBox, 6, 1, 1, 1)
        self.allow_wrong_checkBox = QtWidgets.QCheckBox(self.frame)
        self.allow_wrong_checkBox.setChecked(True)
        self.gridLayout_2.addWidget(self.allow_wrong_checkBox, 10, 0, 1, 1)
        self.okcancel_qpu = QtWidgets.QDialogButtonBox(self.frame)
        self.okcancel_qpu.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel | QtWidgets.QDialogButtonBox.Ok)
        self.gridLayout_2.addWidget(self.okcancel_qpu, 14, 1, 1, 1)
        self.max_questions_spinBox = QtWidgets.QSpinBox(self.frame)
        self.gridLayout_2.addWidget(self.max_questions_spinBox, 3, 1, 1, 1)
        self.label_question_difficulty = QtWidgets.QLabel(self.frame)
        self.gridLayout_2.addWidget(self.label_question_difficulty, 6, 0, 1, 1)
        self.label_max_questions_qpu = QtWidgets.QLabel(self.frame)
        self.gridLayout_2.addWidget(self.label_max_questions_qpu, 3, 0, 1, 1)
        self.challenge_mode_checkBox = QtWidgets.QCheckBox(self.frame)
        self.gridLayout_2.addWidget(self.challenge_mode_checkBox, 10, 1, 1, 1)
        self.exclude_tag_comboBox = QtWidgets.QComboBox(self.frame)
        self.exclude_tag_comboBox.setEnabled(False)
        self.gridLayout_2.addWidget(self.exclude_tag_comboBox, 13, 1, 1, 1)
        self.exclude_tags_checkBox = QtWidgets.QCheckBox(self.frame)
        self.gridLayout_2.addWidget(self.exclude_tags_checkBox, 12, 0, 1, 1)
        self.top_line = QtWidgets.QFrame(self.frame)
        self.top_line.setFrameShape(QtWidgets.QFrame.HLine)
        self.top_line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.gridLayout_2.addWidget(self.top_line, 2, 0, 1, 2)
        self.label_options = QtWidgets.QLabel(self.frame)
        self.gridLayout_2.addWidget(self.label_options, 9, 0, 1, 2)
        self.bottom_line = QtWidgets.QFrame(self.frame)
        self.bottom_line.setFrameShape(QtWidgets.QFrame.HLine)
        self.bottom_line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.gridLayout_2.addWidget(self.bottom_line, 8, 0, 1, 2)
        self.max_questions_hSlider = QtWidgets.QSlider(self.frame)
        self.max_questions_hSlider.setOrientation(QtCore.Qt.Horizontal)
        self.gridLayout_2.addWidget(self.max_questions_hSlider, 4, 1, 1, 1)
        self.mid_line = QtWidgets.QFrame(self.frame)
        self.mid_line.setFrameShape(QtWidgets.QFrame.HLine)
        self.mid_line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.gridLayout_2.addWidget(self.mid_line, 5, 0, 1, 2)
        self.add_questionnaire_to_quiz_button = QtWidgets.QPushButton(self.frame)
        self.gridLayout_2.addWidget(self.add_questionnaire_to_quiz_button, 1, 0, 1, 1)
        self.quiz_comboBox = QtWidgets.QComboBox(self.frame)
        self.gridLayout_2.addWidget(self.quiz_comboBox, 1, 1, 1, 1)
        self.gridLayout.addWidget(self.frame, 0, 0, 1, 1)

        self.exclude_tag_button.setText("Exclude")
        self.label_percentage_max_questions.setText("%")
        self.label_quiz_summary.setText("Quiz Summary\n\n"
                                        "You can build your quizzes by adding one ore more questionnaires.")
        self.label_excluded_tags.setText("")
        self.allow_wrong_checkBox.setText("Allow wrong answers")
        self.label_question_difficulty.setText("Question Difficulty\n(This difficulty or higher)")
        self.label_max_questions_qpu.setText("Max questions in quiz")
        self.challenge_mode_checkBox.setToolTip("<html><head/><body><p align=\"justify\">User gets "
                                                "only certain lives, which are consumed by "
                                                "errors.</p><p align=\"justify\">Get too many mistakes "
                                                "and lose.</p></body></html>")
        self.challenge_mode_checkBox.setText("Challenge Mode")
        self.exclude_tags_checkBox.setText("Exclude tags")
        self.label_options.setText("Options")
        self.add_questionnaire_to_quiz_button.setText("Add Questionnaire")

        # ADDITIONAL INIT
        # -- Add questionnaires to a combo box to let user add multiple ones to a quiz
        for file in os.listdir():
            if file.startswith("Questionnaire"):
                self.quiz_comboBox.addItem(file[14:-5])
        # -- Block dialog functionality until user adds questionnaires to quiz
        self.max_questions_spinBox.setEnabled(False)
        self.max_questions_hSlider.setEnabled(False)
        self.question_difficulty_comboBox.setEnabled(False)
        self.question_difficulty_hSlider.setEnabled(False)

        self.difficulty_dict = {1: "EASY", 3: "MEDIUM",
                                5: "HARD", 7: "EXPERT"}

        # CONNECTIONS
        self.exclude_tags_checkBox.clicked.connect(self.check_exclude_tags)
        self.add_questionnaire_to_quiz_button.clicked.connect(self.add_questionnaire_to_quiz)
        self.max_questions_hSlider.valueChanged.connect(self.max_question_slider_changed)
        self.max_questions_spinBox.valueChanged.connect(self.max_question_spinbox_changed)
        self.question_difficulty_hSlider.valueChanged.connect(self.question_difficulty_slider_changed)
        self.question_difficulty_comboBox.currentTextChanged.connect(self.question_difficulty_combobox_changed)
        self.exclude_tag_button.clicked.connect(self.exclude_this_tag)
        self.okcancel_qpu.accepted.connect(self.create_new_quiz)

        self.exec_()

    def check_exclude_tags(self):
        if self.exclude_tags_checkBox.isChecked():
            self.exclude_tag_button.setEnabled(True)
            self.exclude_tag_comboBox.setEnabled(True)
        else:
            self.exclude_tag_button.setEnabled(False)
            self.exclude_tag_comboBox.setEnabled(False)

    def add_questionnaire_to_quiz(self):
        if self.quiz_comboBox.currentText() != "":
            # FETCHING DATA
            wb = openpyxl.load_workbook(f"Questionnaire_{self.quiz_comboBox.currentText()}.xlsx")
            sheet = wb.active
            # -- List comprehensions fetch data from excel file
            question_list = [cell.value for cell in sheet["A"] if cell.value != "QUESTIONS"]  # EUREKA MOTHER______
            tag_list = [cell.value for cell in sheet["B"] if cell.value != "TAGS" and cell.value is not None]
            difficulty_list = [self.difficulty_dict.get(cell.value) for cell in sheet["H"]
                               if cell.value != "DIFFICULTY"]
            difficulty_list = list(set(difficulty_list))

            # UPDATING UI
            # -- Enable elements once they can be populated from fetched data
            self.max_questions_spinBox.setEnabled(True)
            self.max_questions_hSlider.setEnabled(True)
            self.max_questions_spinBox.setMaximum(len(question_list))
            self.max_questions_hSlider.setMaximum(len(question_list))
            for tag in tag_list:
                # TODO: Manage doubled tags from previous questionnaires, might cause trouble
                self.exclude_tag_comboBox.addItem(tag)

            for diff in sorted(difficulty_list):
                self.question_difficulty_comboBox.addItem(diff)
            self.question_difficulty_comboBox.setEnabled(True)
            self.question_difficulty_hSlider.setEnabled(True)
            self.question_difficulty_hSlider.setMaximum(len(difficulty_list) - 1)

            # -- Remove input questionnaire to avoid duplicates
            questionnaire_list = [self.quiz_comboBox.itemText(i) for i in range(len(self.quiz_comboBox))]
            self.label_quiz_summary.setText(f"{self.label_quiz_summary.text()} \n/{self.quiz_comboBox.currentText()}")
            self.quiz_comboBox.removeItem(questionnaire_list.index(self.quiz_comboBox.currentText()))

            # TESTING
            print(question_list)
            print(difficulty_list)
            print(questionnaire_list)

            wb.close()

    def create_new_quiz(self):
        # CREATING QUIZ
        # -- Ask for quiz title
        title, ok = QtWidgets.QInputDialog.getText(self.window(), "Title", "Select a Title for the new Quiz:",
                                                   QtWidgets.QLineEdit.Normal)
        if title and ok:
            title = title.upper()
            # -- Get un-excluded tags by user
            tag_list = self.label_excluded_tags.text().split("/")
            tag_list.pop(0)
            # -- Open questionnaire and extract filtered questions, add them to another excel file
            questionnaires = (self.label_quiz_summary.text().split("/"))
            questionnaires.pop(0)
            # -- Load first excel file, saves formatting excel files all over again
            first_quiz_wb = openpyxl.load_workbook(f"Questionnaire_{questionnaires[0][:-2]}.xlsx") # TODO: FIX, broken
            first_quiz_wb.save(f"Quiz_{title}.xlsx")
            first_sheet = first_quiz_wb.active
            questionnaires.pop(0)

            # FETCHING DATA
            for questionnaire in questionnaires:
                questionnaire = questionnaire.replace(" \n", "")
                wb = openpyxl.load_workbook(f"Questionnaire_{questionnaire}.xlsx")
                sheet = wb.active
                # -- Delete all rows containing excluded tags
                for cell in sheet["C"]:
                    for tag in tag_list:
                        if tag in cell.value:
                            print(f"Tag {tag} was in {cell.value}")
                            sheet.delete_rows(int(cell.coordinate[1:]), 1)
                # TODO: Limit to max questions in quiz
                # -- Extract relevant info
                question_list = [cell.value for cell in sheet["A"] if cell.value != "QUESTIONS"]
                tag_list = [cell.value for cell in sheet["B"] if cell.value != "TAGS"]
                question_tags_list = [cell.value for cell in sheet["C"] if cell.value != "QUESTION TAGS"]
                answer_list = [cell.value for cell in sheet["D"] if cell.value != "ANSWER"]
                wrong_answer_list_1 = [cell.value for cell in sheet["E"] if cell.value != "WRONG ANSWER 1"]
                wrong_answer_list_2 = [cell.value for cell in sheet["F"] if cell.value != "WRONG ANSWER 2"]
                wrong_answer_list_3 = [cell.value for cell in sheet["G"] if cell.value != "WRONG ANSWER 3"]
                difficulty_list = [cell.value for cell in sheet["H"] if cell.value != "DIFFICULTY"]

                # DUMPING DATA
                # -- Append all row info on new file
                for data in question_list:
                    first_sheet["A" + str(first_sheet.max_row + 1)] = data
                index = first_sheet.max_row - len(question_list) + 1
                for data in tag_list:
                    first_sheet["B" + str(index)] = data
                    index += 1
                index = first_sheet.max_row - len(question_list) + 1
                for data in question_tags_list:
                    first_sheet["C" + str(index)] = data
                    index += 1
                index = first_sheet.max_row - len(question_list) + 1
                for data in answer_list:
                    first_sheet["D" + str(index)] = data
                    index += 1
                index = first_sheet.max_row - len(question_list) + 1
                for data in wrong_answer_list_1:
                    first_sheet["E" + str(index)] = data
                    index += 1
                index = first_sheet.max_row - len(question_list) + 1
                for data in wrong_answer_list_2:
                    first_sheet["F" + str(index)] = data
                    index += 1
                index = first_sheet.max_row - len(question_list) + 1
                for data in wrong_answer_list_3:
                    first_sheet["G" + str(index)] = data
                    index += 1
                index = first_sheet.max_row - len(question_list) + 1
                for data in difficulty_list:
                    first_sheet["H" + str(index)] = data
                    index += 1
                wb.close()
            first_quiz_wb.save(f"Quiz_{title}.xlsx")
            first_quiz_wb.close()

            self.generic_information("Successful Operation", f"Quiz {title} was successfully created!")
            UiQuizMainWindow.__init__()
            CustomQuizPopUp.close(self)

    @staticmethod
    def generic_information(title="Information", text="Details", icon=QMessageBox.Information):
        parent = QtWidgets.QMessageBox()
        parent.setGeometry(800, 300, 500, 500)
        parent.setWindowTitle(title)
        parent.setText(text)
        parent.setIcon(icon)
        parent.exec_()

    @staticmethod
    def generic_error(title="Error", text="Error Details", icon=QMessageBox.Warning):
        parent = QtWidgets.QMessageBox()
        parent.setGeometry(800, 300, 500, 500)
        parent.setWindowTitle(title)
        parent.setText(text)
        parent.setIcon(icon)
        parent.exec_()

    def max_question_slider_changed(self):
        self.max_questions_spinBox.setValue(self.max_questions_hSlider.value())
        if self.max_questions_spinBox.value() != 0:
            percentage = round((self.max_questions_hSlider.value() / self.max_questions_hSlider.maximum()) * 100, 2)
            self.label_percentage_max_questions.setText(f"({self.max_questions_hSlider.value()} / "
                                                        f"{self.max_questions_hSlider.maximum()}) "
                                                        f"{percentage}%")
        else:
            self.label_percentage_max_questions.setText(f"({self.max_questions_hSlider.value()} / "
                                                        f"{self.max_questions_hSlider.maximum()}) "
                                                        f"{0}%")

    def max_question_spinbox_changed(self):
        self.max_questions_hSlider.setValue(self.max_questions_spinBox.value())

    def question_difficulty_slider_changed(self):
        self.question_difficulty_comboBox.setCurrentIndex(self.question_difficulty_hSlider.value())

    def question_difficulty_combobox_changed(self):
        pass

    def exclude_this_tag(self):
        # UPDATE UI
        excluded_tag = self.exclude_tag_comboBox.currentText()
        tag_list = [self.exclude_tag_comboBox.itemText(i) for i in range(self.exclude_tag_comboBox.count())]
        self.exclude_tag_comboBox.removeItem(tag_list.index(excluded_tag))
        self.label_excluded_tags.setText(f"{self.label_excluded_tags.text()}/{excluded_tag}")


class UiQuestionnaireMainWindow(object):
    def __init__(self):
        super().__init__()
        questionnaireMainWindow.setObjectName("QuestionnaireMainWindow")
        questionnaireMainWindow.resize(686, 726)
        questionnaireMainWindow.setGeometry(800, 300, 500, 500)
        self.questionnaire_main_widget = QtWidgets.QWidget(questionnaireMainWindow)
        self.questionnaire_main_widget.setObjectName("questionnaire_main_widget")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.questionnaire_main_widget)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.questionnaire_left_frame = QtWidgets.QFrame(self.questionnaire_main_widget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.questionnaire_left_frame.sizePolicy().hasHeightForWidth())
        self.questionnaire_left_frame.setSizePolicy(sizePolicy)
        self.questionnaire_left_frame.setMinimumSize(QtCore.QSize(180, 0))
        self.questionnaire_left_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.questionnaire_left_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.questionnaire_left_frame.setObjectName("questionnaire_left_frame")
        self.formLayout = QtWidgets.QFormLayout(self.questionnaire_left_frame)
        self.formLayout.setObjectName("formLayout")
        self.new_questionnaire_button = QtWidgets.QPushButton(self.questionnaire_left_frame)
        self.new_questionnaire_button.setMinimumSize(QtCore.QSize(45, 50))
        self.new_questionnaire_button.setStyleSheet("background-color: rgb(0, 170, 255);")
        self.new_questionnaire_button.setObjectName("new_questionnaire_button")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.new_questionnaire_button)
        self.label_new_questionnaire = QtWidgets.QLabel(self.questionnaire_left_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_new_questionnaire.setFont(font)
        self.label_new_questionnaire.setWordWrap(True)
        self.label_new_questionnaire.setObjectName("label_new_questionnaire")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.label_new_questionnaire)
        self.preview_questionnaire_button = QtWidgets.QPushButton(self.questionnaire_left_frame)
        self.preview_questionnaire_button.setMinimumSize(QtCore.QSize(0, 50))
        self.preview_questionnaire_button.setStyleSheet("background-color: rgb(85, 255, 255);")
        self.preview_questionnaire_button.setObjectName("new_question_button")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.preview_questionnaire_button)
        self.label_new_question = QtWidgets.QLabel(self.questionnaire_left_frame)
        self.label_new_question.setWordWrap(True)
        self.label_new_question.setObjectName("label_new_question")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.label_new_question)
        self.label_project_nav = QtWidgets.QLabel(self.questionnaire_left_frame)
        self.label_project_nav.setWordWrap(True)
        self.label_project_nav.setObjectName("label_project_nav")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.SpanningRole, self.label_project_nav)
        self.main_list_widget = QtWidgets.QListWidget(self.questionnaire_left_frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.main_list_widget.sizePolicy().hasHeightForWidth())
        self.main_list_widget.setSizePolicy(sizePolicy)
        self.main_list_widget.setMinimumSize(QtCore.QSize(81, 0))
        self.main_list_widget.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.main_list_widget.setLineWidth(3)
        self.main_list_widget.setObjectName("main_list_widget")
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.SpanningRole, self.main_list_widget)
        self.select_current_questionnaire = QtWidgets.QPushButton(self.questionnaire_left_frame)
        self.select_current_questionnaire.setObjectName("change_quiz_button")
        self.formLayout.setWidget(4, QtWidgets.QFormLayout.SpanningRole, self.select_current_questionnaire)
        # UI CHANGE ----------------------------------------------------------------------------------------------
        self.modify_questionnaire_title_button = QtWidgets.QPushButton()
        self.modify_questionnaire_title_button.setText("Modify questionnaire Title")
        self.formLayout.setWidget(5, QtWidgets.QFormLayout.SpanningRole, self.modify_questionnaire_title_button)
        self.delete_questionnaire_button = QtWidgets.QPushButton()
        self.delete_questionnaire_button.setText("Delete questionnaire")
        self.formLayout.setWidget(6, QtWidgets.QFormLayout.SpanningRole, self.delete_questionnaire_button)
        self.modify_question_button = QtWidgets.QPushButton()
        self.modify_question_button.setText("Modify question")
        self.formLayout.setWidget(7, QtWidgets.QFormLayout.SpanningRole, self.modify_question_button)
        self.delete_question_button = QtWidgets.QPushButton()
        self.delete_question_button.setText("Delete question")
        self.formLayout.setWidget(8, QtWidgets.QFormLayout.SpanningRole, self.delete_question_button)
        # UI CHANGE FINISH ---------------------------------------------------------------------------------------
        self.go_to_quizbutton = QtWidgets.QPushButton(self.questionnaire_left_frame)
        self.go_to_quizbutton.setObjectName("go_to_questionnaire_button")
        self.formLayout.setWidget(10, QtWidgets.QFormLayout.SpanningRole, self.go_to_quizbutton)
        spacerItem4 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.formLayout.setItem(9, QtWidgets.QFormLayout.SpanningRole, spacerItem4)
        self.gridLayout_2.addWidget(self.questionnaire_left_frame, 0, 0, 1, 1)
        self.questionnaire_right_frame = QtWidgets.QFrame(self.questionnaire_main_widget)
        self.questionnaire_right_frame.setEnabled(True)
        self.questionnaire_right_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.questionnaire_right_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.questionnaire_right_frame.setObjectName("questionnaire_right_frame")
        self.gridLayout = QtWidgets.QGridLayout(self.questionnaire_right_frame)
        self.gridLayout.setObjectName("gridLayout")
        self.wrong_answer_text_edit_3 = QtWidgets.QTextEdit(self.questionnaire_right_frame)
        self.wrong_answer_text_edit_3.setEnabled(False)
        self.wrong_answer_text_edit_3.setObjectName("alternate_answer_text_edit_3")
        self.gridLayout.addWidget(self.wrong_answer_text_edit_3, 15, 3, 1, 2)
        self.medium_check_box = QtWidgets.QCheckBox(self.questionnaire_right_frame)
        self.medium_check_box.setAutoExclusive(True)
        self.medium_check_box.setObjectName("medium_check_box")
        self.gridLayout.addWidget(self.medium_check_box, 17, 2, 1, 1)
        self.label_number_of_questions = QtWidgets.QLabel(self.questionnaire_right_frame)
        self.label_number_of_questions.setObjectName("label_number_of_questions")
        self.gridLayout.addWidget(self.label_number_of_questions, 3, 0, 1, 1)
        self.remove_tag_button = QtWidgets.QPushButton(self.questionnaire_right_frame)
        self.remove_tag_button.setObjectName("remove_tag_button")
        self.gridLayout.addWidget(self.remove_tag_button, 10, 4, 1, 1)
        self.add_tag_to_question_button = QtWidgets.QPushButton(self.questionnaire_right_frame)
        self.add_tag_to_question_button.setStyleSheet("background-color: rgb(0, 0, 0);\n"
                                                      "color: rgb(255, 255, 255);")
        self.add_tag_to_question_button.setObjectName("add_tag_to_question_button")
        self.gridLayout.addWidget(self.add_tag_to_question_button, 9, 4, 1, 1)
        self.label_tags = QtWidgets.QLabel(self.questionnaire_right_frame)
        self.label_tags.setObjectName("label_tags")
        self.gridLayout.addWidget(self.label_tags, 9, 0, 1, 1)
        self.answer_text_edit = QtWidgets.QTextEdit(self.questionnaire_right_frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.answer_text_edit.sizePolicy().hasHeightForWidth())
        self.answer_text_edit.setSizePolicy(sizePolicy)
        self.answer_text_edit.setMinimumSize(QtCore.QSize(0, 150))
        self.answer_text_edit.setObjectName("answer_text_edit")
        self.gridLayout.addWidget(self.answer_text_edit, 12, 0, 1, 5)
        self.add_question_button = QtWidgets.QPushButton(self.questionnaire_right_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.add_question_button.setFont(font)
        self.add_question_button.setStyleSheet("background-color: rgb(0, 255, 0);")
        self.add_question_button.setCheckable(False)
        self.add_question_button.setAutoDefault(False)
        self.add_question_button.setDefault(False)
        self.add_question_button.setFlat(False)
        self.add_question_button.setObjectName("add_question_button")
        self.gridLayout.addWidget(self.add_question_button, 18, 0, 1, 4)
        self.add_tag_button = QtWidgets.QPushButton(self.questionnaire_right_frame)
        self.add_tag_button.setObjectName("add_tag_button")
        self.gridLayout.addWidget(self.add_tag_button, 10, 1, 1, 2)
        spacerItem5 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem5, 5, 0, 1, 1)
        self.tag_combo_box = QtWidgets.QComboBox(self.questionnaire_right_frame)
        self.tag_combo_box.setFixedHeight(23)  # UI SIZE CHANGED
        self.tag_combo_box.setStyleSheet("background-color: rgb(0, 0, 0);\n"
                                         "color: rgb(255, 255, 255);")
        self.tag_combo_box.setObjectName("tag_combo_box")
        self.gridLayout.addWidget(self.tag_combo_box, 9, 1, 1, 3)
        self.easy_check_box = QtWidgets.QCheckBox(self.questionnaire_right_frame)
        self.easy_check_box.setChecked(True)
        self.easy_check_box.setAutoExclusive(True)
        self.easy_check_box.setObjectName("easy_check_box")
        self.gridLayout.addWidget(self.easy_check_box, 17, 0, 1, 1)
        self.label_average_difficulty = QtWidgets.QLabel(self.questionnaire_right_frame)
        self.label_average_difficulty.setObjectName("label_average_difficulty")
        self.gridLayout.addWidget(self.label_average_difficulty, 3, 3, 1, 1)
        self.label_project_title_questionnaire = QtWidgets.QLabel(self.questionnaire_right_frame)
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.label_project_title_questionnaire.setFont(font)
        self.label_project_title_questionnaire.setObjectName("label_project_title_questionnaire")
        self.gridLayout.addWidget(self.label_project_title_questionnaire, 0, 0, 1, 3)
        self.wrong_answer_text_edit_2 = QtWidgets.QTextEdit(self.questionnaire_right_frame)
        self.wrong_answer_text_edit_2.setEnabled(False)
        self.wrong_answer_text_edit_2.setObjectName("alternate_answer_text_edit_2")
        self.gridLayout.addWidget(self.wrong_answer_text_edit_2, 15, 0, 1, 3)
        self.label_remove_tag = QtWidgets.QLabel(self.questionnaire_right_frame)
        self.label_remove_tag.setObjectName("label_remove_tag")
        self.gridLayout.addWidget(self.label_remove_tag, 10, 3, 1, 1)
        self.possible_wrong_answers_check_box = QtWidgets.QCheckBox(self.questionnaire_right_frame)
        self.possible_wrong_answers_check_box.setMinimumSize(QtCore.QSize(0, 32))
        self.possible_wrong_answers_check_box.setIconSize(QtCore.QSize(16, 32))
        self.possible_wrong_answers_check_box.setObjectName("possible_wrong_answers_check_box")
        self.gridLayout.addWidget(self.possible_wrong_answers_check_box, 14, 0, 1, 3)
        spacerItem6 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout.addItem(spacerItem6, 6, 0, 1, 1)
        self.question_line_edit = QtWidgets.QLineEdit(self.questionnaire_right_frame)
        self.question_line_edit.setMinimumSize(QtCore.QSize(0, 40))
        self.question_line_edit.setText("")
        self.question_line_edit.setObjectName("question_line_edit")
        self.gridLayout.addWidget(self.question_line_edit, 8, 0, 1, 5)
        self.label_questionnaire_details = QtWidgets.QLabel(self.questionnaire_right_frame)
        self.label_questionnaire_details.setObjectName("label_questionnaire_details")
        self.gridLayout.addWidget(self.label_questionnaire_details, 2, 0, 1, 5)
        self.label_add_tag = QtWidgets.QLabel(self.questionnaire_right_frame)
        self.label_add_tag.setObjectName("label_add_tag")
        self.gridLayout.addWidget(self.label_add_tag, 10, 0, 1, 1)
        self.cancel_button = QtWidgets.QPushButton(self.questionnaire_right_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.cancel_button.setFont(font)
        self.cancel_button.setStyleSheet("background-color: rgb(255, 0, 0);")
        self.cancel_button.setObjectName("cancel_button")
        self.gridLayout.addWidget(self.cancel_button, 18, 4, 1, 1)
        self.wrong_answer_text_edit_1 = QtWidgets.QTextEdit(self.questionnaire_right_frame)
        self.wrong_answer_text_edit_1.setEnabled(False)
        self.wrong_answer_text_edit_1.setObjectName("alternate_answer_text_edit_1")
        self.gridLayout.addWidget(self.wrong_answer_text_edit_1, 14, 3, 1, 2)
        self.label_answer = QtWidgets.QLabel(self.questionnaire_right_frame)
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_answer.setFont(font)
        self.label_answer.setObjectName("label_answer")
        self.gridLayout.addWidget(self.label_answer, 11, 0, 1, 1)
        self.label_question = QtWidgets.QLabel(self.questionnaire_right_frame)
        self.label_question.setObjectName("label_question")
        self.gridLayout.addWidget(self.label_question, 7, 0, 1, 1)
        self.expert_check_box = QtWidgets.QCheckBox(self.questionnaire_right_frame)
        self.expert_check_box.setAutoExclusive(True)
        self.expert_check_box.setObjectName("expert_check_box")
        self.gridLayout.addWidget(self.expert_check_box, 17, 4, 1, 1)
        self.hard_check_box = QtWidgets.QCheckBox(self.questionnaire_right_frame)
        self.hard_check_box.setAutoExclusive(True)
        self.hard_check_box.setObjectName("hard_check_box")
        self.gridLayout.addWidget(self.hard_check_box, 17, 3, 1, 1)
        self.gridLayout_2.addWidget(self.questionnaire_right_frame, 0, 2, 1, 1)
        self.line = QtWidgets.QFrame(self.questionnaire_main_widget)
        self.line.setFrameShape(QtWidgets.QFrame.VLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.gridLayout_2.addWidget(self.line, 0, 1, 1, 1)
        questionnaireMainWindow.setCentralWidget(self.questionnaire_main_widget)
        self.questionnaire_menu_bar = QtWidgets.QMenuBar(questionnaireMainWindow)
        self.questionnaire_menu_bar.setGeometry(QtCore.QRect(0, 0, 686, 21))
        self.questionnaire_menu_bar.setObjectName("questionnaire_menu_bar")
        questionnaireMainWindow.setMenuBar(self.questionnaire_menu_bar)
        self.questionnaire_status_bar = QtWidgets.QStatusBar(questionnaireMainWindow)
        self.questionnaire_status_bar.setObjectName("questionnaire_status_bar")
        questionnaireMainWindow.setStatusBar(self.questionnaire_status_bar)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        self.textify(questionnaireMainWindow)
        self.possible_wrong_answers_check_box.toggled['bool'].connect(self.wrong_answer_text_edit_1.setEnabled)
        self.possible_wrong_answers_check_box.toggled['bool'].connect(self.wrong_answer_text_edit_2.setEnabled)
        self.possible_wrong_answers_check_box.toggled['bool'].connect(self.wrong_answer_text_edit_3.setEnabled)
        QtCore.QMetaObject.connectSlotsByName(questionnaireMainWindow)

        # CHILD WINDOW VARIABLE
        self.window = QtWidgets.QWidget()
        self.window.setGeometry(900, 400, 500, 500)

        # UI Additions
        self.menuOpciones = QtWidgets.QMenu(self.questionnaire_menu_bar)
        self.menuOpciones.setTitle("Opciones")
        self.label_added_tag_one = QtWidgets.QLabel(self.questionnaire_right_frame)
        self.label_added_tag_one.setText("")
        self.gridLayout.addWidget(self.label_added_tag_one, 7, 3, 1, 1)
        self.label_added_tag_two = QtWidgets.QLabel(self.questionnaire_right_frame)
        self.label_added_tag_two.setText("")
        self.gridLayout.addWidget(self.label_added_tag_two, 7, 4, 1, 1)
        self.mod_add_question_button = QtWidgets.QPushButton(self.questionnaire_right_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.mod_add_question_button.setFont(font)
        self.mod_add_question_button.setStyleSheet("background-color: rgb(0, 255, 0);")
        self.mod_add_question_button.setCheckable(False)
        self.mod_add_question_button.setAutoDefault(False)
        self.mod_add_question_button.setDefault(False)
        self.mod_add_question_button.setFlat(False)
        self.gridLayout.addWidget(self.mod_add_question_button, 18, 0, 1, 4)
        self.mod_add_question_button.setText("(-) Modify Question!")
        self.mod_add_question_button.hide()
        self.questionnaire_left_frame.setFixedWidth(180)
        self.go_to_quizbutton.setEnabled(False)
        # -- Placeholder texts
        self.question_line_edit.setPlaceholderText("Your question goes here:")
        self.answer_text_edit.setPlaceholderText("Your answer goes here:")
        self.wrong_answer_text_edit_1.setPlaceholderText("Your wrong answer goes here: (Trickery)")
        self.wrong_answer_text_edit_2.setPlaceholderText("Your wrong answer goes here: (Banditry)")
        self.wrong_answer_text_edit_3.setPlaceholderText("Your wrong answer goes here: (Deception)")
        # -- Tooltips
        self.new_questionnaire_button.setToolTip("You can add new Questionnaires here!")
        self.select_current_questionnaire.setToolTip("Select a Questionnaire from above to set it as active")

        # TAB ORDER
        MainWindow.setTabOrder(self.question_line_edit, self.answer_text_edit)

        # ADDITIONAL INIT # -- Sets regular path to project files
        self.path = os.getcwd()
        # -- Add current questionnaires to main list widget
        for file in os.listdir():
            if file.startswith("Questionnaire"):
                self.main_list_widget.addItem(file[14:-5])
        if self.main_list_widget.count() == 0:
            self.questionnaire_right_frame.setEnabled(False)
        else:
            self.set_active_questionnaire(self.main_list_widget.count() - 1)
        if self.label_added_tag_one.text() != "" and self.label_added_tag_two.text() != "":
            self.add_tag_to_question_button.setText("Remove tags >")
        self.difficulty_dict = {1: "Easy", 3: "Medium", 5: "Hard", 7: "Expert"}

        # CONNECTIONS
        self.go_to_quizbutton.clicked.connect(self.switch_to_quiz)
        self.new_questionnaire_button.clicked.connect(self.new_questionnaire)
        self.modify_questionnaire_title_button.clicked.connect(self.modify_title)
        self.delete_questionnaire_button.clicked.connect(self.delete_questionnaire)
        self.add_question_button.clicked.connect(self.add_question)
        self.select_current_questionnaire.clicked.connect(self.select_current)
        self.delete_question_button.clicked.connect(self.delete_question)
        self.add_tag_button.clicked.connect(self.add_tags)
        self.remove_tag_button.clicked.connect(self.remove_tags)
        self.add_tag_to_question_button.clicked.connect(self.add_tag_to_question)
        self.modify_question_button.clicked.connect(self.modify_question)
        self.mod_add_question_button.clicked.connect(self.add_mod_question)
        self.cancel_button.clicked.connect(self.questionnaire_cancel)
        self.preview_questionnaire_button.clicked.connect(self.preview_questionnaire)

    def textify(self, questionnaireMainWindow):
        _translate = QtCore.QCoreApplication.translate
        questionnaireMainWindow.setWindowTitle(_translate("QuestionnaireMainWindow", "Questionnaire"))
        self.new_questionnaire_button.setText(_translate("QuestionnaireMainWindow", "(🞤)"))
        self.label_new_questionnaire.setText(_translate("QuestionnaireMainWindow", "New questionnaire \n"
                                                                                   "(This goes first!)"))
        self.preview_questionnaire_button.setText(_translate("QuestionnaireMainWindow", "(🔎)"))
        self.label_new_question.setText(_translate("QuestionnaireMainWindow", "Preview current questionnaire"))
        self.label_project_nav.setText(
            _translate("QuestionnaireMainWindow", "Click here to navigate through projects:"))
        self.select_current_questionnaire.setText(_translate("QuestionnaireMainWindow", "Select current questionnaire"))
        self.go_to_quizbutton.setText(_translate("QuestionnaireMainWindow", "Lets take a"
                                                                            " Quiz! \n(Needs 4 questions at least)"))
        self.medium_check_box.setText(_translate("QuestionnaireMainWindow", "Medium (+3 points)"))
        self.label_number_of_questions.setText(_translate("QuestionnaireMainWindow", "# of questions:"))
        self.remove_tag_button.setText(_translate("QuestionnaireMainWindow", "(-)"))
        self.add_tag_to_question_button.setText(_translate("QuestionnaireMainWindow", "Add tag >"))
        self.label_tags.setText(_translate("QuestionnaireMainWindow", "Tags"))
        self.add_question_button.setText(_translate("QuestionnaireMainWindow", "(✓) Add Question!"))
        self.add_tag_button.setText(_translate("QuestionnaireMainWindow", "(+)"))
        self.easy_check_box.setText(_translate("QuestionnaireMainWindow", "Easy (+1 point)"))
        self.label_average_difficulty.setText(_translate("QuestionnaireMainWindow", "Average Difficulty:"))
        self.label_project_title_questionnaire.setText(_translate("QuestionnaireMainWindow", "Questionnaire Title"))
        self.label_remove_tag.setText(_translate("QuestionnaireMainWindow", "Remove Tag"))
        self.possible_wrong_answers_check_box.setText(_translate("QuestionnaireMainWindow", "Possible Wrong Answers"))
        self.label_questionnaire_details.setText(_translate("QuestionnaireMainWindow", "Questionnaire Details:"))
        self.label_add_tag.setText(_translate("QuestionnaireMainWindow", "Add Tag"))
        self.cancel_button.setText(_translate("QuestionnaireMainWindow", "(X) Cancel"))
        self.label_answer.setText(_translate("QuestionnaireMainWindow", "Answer"))
        self.label_question.setText(_translate("QuestionnaireMainWindow", "Question"))
        self.expert_check_box.setText(_translate("QuestionnaireMainWindow", "Expert (+7 points)"))
        self.hard_check_box.setText(_translate("QuestionnaireMainWindow", "Hard (+5 points)"))

    def set_active_questionnaire(self, index=0):
        # FETCHING DATA
        wb = openpyxl.load_workbook(f"Questionnaire_{self.main_list_widget.item(index).text()}.xlsx")
        sheet = wb.active
        # -- Fetching questions column in excel file, adding it to a list
        question_list = [cell.value for cell in sheet["A"] if cell.value != "QUESTIONS" and cell.value is not None]

        # UPDATE UI -- Updating labels and status bar to reflect fetched data
        self.label_number_of_questions.setText(f"""# of Questions: {len(question_list)}""")
        self.label_question.setText(f"""Question {len(question_list) + 1}:""")
        self.questionnaire_status_bar.showMessage(f"Active questionnaire: "
                                                  f"{self.main_list_widget.item(index).text()}")
        self.label_project_title_questionnaire.setText(self.main_list_widget.item(index).text())
        # -- Clearing combo box from possible other questionnaires
        self.tag_combo_box.clear()

        # FETCHING DATA AGAIN -- Fetching tags from excel file, adding them to a list
        for cell in sheet["B"]:  # TAGS
            if cell.value != "TAGS" and cell.value is not None:
                self.tag_combo_box.addItem(cell.value)
        # -- Fetching current question, enabling quiz button or not
        current_question = int(self.label_question.text().replace(":", "")[9:])
        if current_question >= 4:
            self.go_to_quizbutton.setEnabled(True)
        # -- Set question tag info into UI
        if sheet["C" + str(current_question + 1)].value is not None:
            tags = sheet["C" + str(current_question + 1)].value.split("//")
            self.label_added_tag_one.setText(tags[0])
            try:
                self.label_added_tag_two.setText(tags[1])
            except IndexError:
                pass
        else:
            self.label_added_tag_one.setText("")
            self.label_added_tag_two.setText("")
        # -- Fetching difficulty values
        if sheet["H2"].value != "" and sheet["H2"].value is not None:  # DIFFICULTY
            diff_average = 0
            count = 0
            for cell in sheet["H"]:
                if cell.value != "DIFFICULTY" and cell.value is not None:
                    diff_average = diff_average + cell.value
                    count += 1
            average = round(diff_average / count, 2)
            self.label_average_difficulty.setText(f"Average Difficulty: {str(average)}")
        # -- If there are no difficulty values, then there are no questions, therefore add default UI message
        else:
            self.label_average_difficulty.setText("Average Difficulty: Pending...")
        # -- Save and close questionnaire excel file
        wb.save(f"Questionnaire_{self.main_list_widget.item(index).text()}.xlsx")
        wb.close()

    def get_active_questionnaire_index(self):
        # MANAGE DATA
        # -- Get all items in main list widget, add them to a list and the find the index of the one marked as active
        items = []
        for i in range(self.main_list_widget.count()):
            items.append(self.main_list_widget.item(i).text())
        current_questionnaire = items.index(self.label_project_title_questionnaire.text())
        return current_questionnaire

    def new_questionnaire(self):
        quiz_title, ok = QtWidgets.QInputDialog.getText(self.window, "Create a Quiz", "Select a title for your "
                                                                                      "questionnaire: ",
                                                        QtWidgets.QLineEdit.Normal)
        if ok and quiz_title:
            wb = xlsxwriter.Workbook(f"Questionnaire_{quiz_title.upper()}.xlsx")
            wb.close()
            open_workbook = openpyxl.load_workbook(f"Questionnaire_{quiz_title.upper()}.xlsx")
            sheet = open_workbook.active
            # Formatting
            sheet["A1"] = "QUESTIONS"
            sheet["B1"] = "TAGS"
            sheet["C1"] = "QUESTION TAGS"
            sheet["D1"] = "ANSWER"
            sheet["E1"] = "WRONG ANSWER 1"
            sheet["F1"] = "WRONG ANSWER 2"
            sheet["G1"] = "WRONG ANSWER 3"
            sheet["H1"] = "DIFFICULTY"
            # Column width
            sheet.column_dimensions["A"].width = 30
            sheet.column_dimensions["B"].width = 24
            sheet.column_dimensions["C"].width = 24
            sheet.column_dimensions["D"].width = 30
            sheet.column_dimensions["E"].width = 30
            sheet.column_dimensions["F"].width = 30
            sheet.column_dimensions["G"].width = 30
            sheet.column_dimensions["H"].width = 15
            # Filter and freeze
            sheet.auto_filter.ref = "A1:H1"
            sheet.freeze_panes = "A2"
            open_workbook.save(f"Questionnaire_{quiz_title.upper()}.xlsx")
            open_workbook.close()
            questionnaire_ui.__init__()

    def select_current(self):
        if len(self.main_list_widget.selectedItems()) > 0:  # Checks if there is any item selected
            item_list = [self.main_list_widget.item(i).text() for i in range(len(self.main_list_widget))]
            self.set_active_questionnaire(item_list.index(self.main_list_widget.selectedItems()[0].text()))
            if self.label_added_tag_one.text() == "" or self.label_added_tag_two.text() == "":
                self.add_tag_to_question_button.setText("Add tag >")
            elif self.label_added_tag_one.text() != "" or self.label_added_tag_two.text() != "":
                self.add_tag_to_question_button.setText("Remove tags >")
        else:
            self.generic_error("No questionnaire selected",
                               "Please select a questionnaire to switch to from the list shown above.")

    def modify_title(self):
        if len(self.main_list_widget.selectedItems()) > 0:
            new_title, ok = QtWidgets.QInputDialog.getText(self.window, "Change questionnaire Title",
                                                           'Select new questionnaire title',
                                                           QtWidgets.QLineEdit.Normal)
            if new_title and ok:
                old_title = f"Questionnaire_{self.main_list_widget.selectedItems()[0].text()}.xlsx"
                new_title = f"Questionnaire_{new_title.upper()}.xlsx"
                if old_title.lower() == new_title.lower():
                    self.generic_error("Identical Titles",
                                       "The two selected titles were identical, operation was cancelled.")
                    return None
                wb = openpyxl.load_workbook(old_title)
                wb.save(new_title)
                wb.close()
                os.remove(old_title)
                questionnaire_ui.__init__()
        else:
            self.generic_error("No title selected", "Select a title from the list above to modify its title")

    def delete_questionnaire(self):
        confirmation = QtWidgets.QMessageBox.question(self.window, "Confirmation",
                                                      "Are you sure you want to permanently delete this questionnaire:")
        if confirmation == QMessageBox.Yes:
            try:
                deleted_quiz = f"Questionnaire_{self.main_list_widget.selectedItems()[0].text()}.xlsx"
                os.remove(deleted_quiz)
                questionnaire_ui.__init__()
            except IndexError:
                self.generic_error("No questionnaire selected", "Please select a questionnaire from the list above to "
                                                                "delete it")
        else:
            self.generic_information("", "Questionnaire was not deleted.")

    def add_question(self):
        # Checks if there are no questionnaire currently created, presents option to create one now
        if self.main_list_widget.count() == 0:  # Counts the items in ItemListWidget
            ask = QtWidgets.QMessageBox.question(self.window, "No quiz created", "It appears you have not created "
                                                                                 "a quiz yet to assign this question\n"
                                                                                 "\nWould you like to create one now?")
            if ask == QMessageBox.Yes:
                self.new_questionnaire()
        # Checks for incomplete user input, manages errors
        else:
            if self.question_line_edit.text() == "":
                self.generic_error("Empty question", "Please add text to your question.")
            elif self.answer_text_edit.toPlainText() == "":
                self.generic_error("Empty answer", "Please add a correct answer to your question.")

            # HANDLES OPTIONAL MULTIPLE WRONG ANSWERS MODE
            elif self.possible_wrong_answers_check_box.isChecked():
                if self.wrong_answer_text_edit_1.toPlainText() == "":
                    self.generic_error("Empty answer", "Please add another wrong answer to your question.\n\n"
                                                       "Otherwise, uncheck possible wrong answer mode.")
                elif self.wrong_answer_text_edit_2.toPlainText() == "":
                    self.generic_error("Empty answer", "Please add another wrong answer to your question.\n\n"
                                                       "Otherwise, uncheck possible wrong answer mode.")
                elif self.wrong_answer_text_edit_3.toPlainText() == "":
                    self.generic_error("Empty answer", "Please add another wrong answer to your question.\n\n"
                                                       "Otherwise, uncheck possible wrong answer mode.")
                # END OF EMPTY FIELD HANDLING

                # START OF CORRECT INPUT HANDLING
                else:  # If all fields are correctly filled, adds question data to file
                    wb = openpyxl.load_workbook(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
                    sheet = wb.active
                    current_question = int(self.label_question.text().replace(":", "")[9:])
                    sheet["A" + str(current_question + 1)] = self.question_line_edit.text().capitalize()  # QUESTION
                    sheet["D" + str(current_question + 1)] = self.answer_text_edit.toPlainText().capitalize()  # ANSWER
                    if self.possible_wrong_answers_check_box.isChecked():  # POSSIBLE WRONG ANSWERS
                        sheet[
                            "E" + str(current_question + 1)] = self.wrong_answer_text_edit_1.toPlainText().capitalize()
                        sheet[
                            "F" + str(current_question + 1)] = self.wrong_answer_text_edit_2.toPlainText().capitalize()
                        sheet[
                            "G" + str(current_question + 1)] = self.wrong_answer_text_edit_3.toPlainText().capitalize()
                    if self.easy_check_box.isChecked():  # DIFFICULTY
                        difficulty = 1
                        sheet["H" + str(current_question + 1)] = difficulty
                    elif self.medium_check_box.isChecked():
                        difficulty = 3
                        sheet["H" + str(current_question + 1)] = difficulty
                    elif self.hard_check_box.isChecked():
                        difficulty = 5
                        sheet["H" + str(current_question + 1)] = difficulty
                    else:
                        difficulty = 7
                        sheet["H" + str(current_question + 1)] = difficulty

                    wb.save(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
                    wb.close()
                    self.generic_information("Successfully added",
                                             f"Question {current_question} was added correctly.\n\n"
                                             f"Question: {self.question_line_edit.text().capitalize()}\n\n"
                                             f"Answer: {self.answer_text_edit.toPlainText().capitalize()}\n\n"
                                             f"Tags: {self.label_added_tag_one.text()} / {self.label_added_tag_two.text()}"
                                             f"\n\nDifficulty: {self.difficulty_dict.get(difficulty)}")
                    self.questionnaire_cancel()

                    items = []
                    for i in range(self.main_list_widget.count()):
                        items.append(self.main_list_widget.item(i).text())
                    current_questionnaire = items.index(self.label_project_title_questionnaire.text())
                    self.set_active_questionnaire(current_questionnaire)

            # HANDLES REGULAR ANSWER MODE
            else:
                wb = openpyxl.load_workbook(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
                sheet = wb.active
                current_question = int(self.label_question.text().replace(":", "")[9:])
                sheet["A" + str(current_question + 1)] = self.question_line_edit.text().capitalize()  # QUESTION
                sheet["D" + str(current_question + 1)] = self.answer_text_edit.toPlainText().capitalize()  # ANSWER

                if self.easy_check_box.isChecked():  # DIFFICULTY
                    difficulty = 1
                    sheet["H" + str(current_question + 1)] = difficulty
                elif self.medium_check_box.isChecked():
                    difficulty = 3
                    sheet["H" + str(current_question + 1)] = difficulty
                elif self.hard_check_box.isChecked():
                    difficulty = 5
                    sheet["H" + str(current_question + 1)] = difficulty
                else:
                    difficulty = 7
                    sheet["H" + str(current_question + 1)] = difficulty

                wb.save(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
                wb.close()
                self.generic_information("Successfully added",
                                         f"Question {current_question} was added correctly.\n\n"
                                         f"Question: {self.question_line_edit.text().capitalize()}\n\n"
                                         f"Answer: {self.answer_text_edit.toPlainText().capitalize()}\n\n"
                                         f"Tags: {self.label_added_tag_one.text()} / {self.label_added_tag_two.text()}"
                                         f"\n\nDifficulty: {self.difficulty_dict.get(difficulty)}")
                self.questionnaire_cancel()

                items = []
                for i in range(self.main_list_widget.count()):
                    items.append(self.main_list_widget.item(i).text())
                current_questionnaire = items.index(self.label_project_title_questionnaire.text())
                self.set_active_questionnaire(current_questionnaire)

    def questionnaire_cancel(self):
        self.question_line_edit.clear()
        self.answer_text_edit.clear()
        self.wrong_answer_text_edit_1.clear()
        self.wrong_answer_text_edit_2.clear()
        self.wrong_answer_text_edit_3.clear()
        self.label_added_tag_one.setText("")
        self.label_added_tag_two.setText("")
        self.add_tag_to_question_button.setText("Add tag >")

    def modify_question(self):
        # FETCHING DATA
        if self.label_project_title_questionnaire.text() != "Questionnaire Title":
            wb = openpyxl.load_workbook(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
            sheet = wb.active
            question_list = []
            for cell in sheet["A"]:
                if cell.value != "QUESTIONS" and cell.value is not None:
                    question_list.append(cell.value)
            question, ok = QtWidgets.QInputDialog.getItem(self.window, "Modify Question",
                                                          "Select a question to modify:",
                                                          question_list, 0, False)
            if question and ok:
                # Depending on selection, bring the question parameters into the UI for modification
                # -- Set Question
                self.question_line_edit.setText(sheet["A" + str(question_list.index(question) + 2)].value)
                # -- Set Answer
                answer = sheet["D" + str(question_list.index(question) + 2)].value
                self.answer_text_edit.setText(answer)
                # -- Set Possible Wrong answers, if there are any
                if sheet["D" + str(question_list.index(question) + 2)].value is not None:
                    self.wrong_answer_text_edit_1.setText(sheet["E" + str(question_list.index(question) + 2)].value)
                    self.wrong_answer_text_edit_2.setText(sheet["F" + str(question_list.index(question) + 2)].value)
                    self.wrong_answer_text_edit_3.setText(sheet["G" + str(question_list.index(question) + 2)].value)
                    self.possible_wrong_answers_check_box.setChecked(True)

                # DONE: Don't add question, but modify the one selected
                self.add_question_button.hide()
                self.mod_add_question_button.show()
                self.label_question.setText(f"Modifying Question {str(question_list.index(question) + 1)}:")

            wb.save(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
            wb.close()

    def add_mod_question(self):
        if self.question_line_edit.text() == "":
            self.generic_error("Empty question", "Please add text to your question.")
        elif self.answer_text_edit.toPlainText() == "":
            self.generic_error("Empty answer", "Please add a correct answer to your question.")
        elif self.possible_wrong_answers_check_box.isChecked():
            if self.wrong_answer_text_edit_1.toPlainText() == "":
                self.generic_error("Empty answer", "Please add another wrong answer to your question.\n\n"
                                                   "Otherwise, uncheck possible wrong answer mode.")
            elif self.wrong_answer_text_edit_2.toPlainText() == "":
                self.generic_error("Empty answer", "Please add another wrong answer to your question.\n\n"
                                                   "Otherwise, uncheck possible wrong answer mode.")
            elif self.wrong_answer_text_edit_3.toPlainText() == "":
                self.generic_error("Empty answer", "Please add another wrong answer to your question.\n\n"
                                                   "Otherwise, uncheck possible wrong answer mode.")
            else:  # If all fields are correctly filled, adds question data to file
                wb = openpyxl.load_workbook(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
                sheet = wb.active
                current_question = int(self.label_question.text().replace(":", "")[9:])
                sheet["A" + str(current_question + 1)] = self.question_line_edit.text().capitalize()  # QUESTION
                sheet["D" + str(current_question + 1)] = self.answer_text_edit.toPlainText().capitalize()  # ANSWER
                if self.possible_wrong_answers_check_box.isChecked():  # POSSIBLE WRONG ANSWERS
                    sheet[
                        "E" + str(current_question + 1)] = self.wrong_answer_text_edit_1.toPlainText().capitalize()
                    sheet[
                        "F" + str(current_question + 1)] = self.wrong_answer_text_edit_2.toPlainText().capitalize()
                    sheet[
                        "G" + str(current_question + 1)] = self.wrong_answer_text_edit_3.toPlainText().capitalize()
                if self.easy_check_box.isChecked():  # DIFFICULTY
                    difficulty = 1
                    sheet["H" + str(current_question + 1)] = difficulty
                elif self.medium_check_box.isChecked():
                    difficulty = 3
                    sheet["H" + str(current_question + 1)] = difficulty
                elif self.hard_check_box.isChecked():
                    difficulty = 5
                    sheet["H" + str(current_question + 1)] = difficulty
                else:
                    difficulty = 7
                    sheet["H" + str(current_question + 1)] = difficulty

                wb.save(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
                wb.close()
                self.generic_information("Successfully added",
                                         f"Question {current_question} was added correctly.\n\n"
                                         f"Question: {self.question_line_edit.text().capitalize()}\n\n"
                                         f"Answer: {self.answer_text_edit.toPlainText().capitalize()}\n\n"
                                         f"Tags: {self.label_added_tag_one.text()} / {self.label_added_tag_two.text()}"
                                         f"\n\nDifficulty: {self.difficulty_dict.get(difficulty)}")
                self.question_line_edit.clear()
                self.answer_text_edit.clear()
                self.wrong_answer_text_edit_1.clear()
                self.wrong_answer_text_edit_2.clear()
                self.wrong_answer_text_edit_3.clear()
                self.label_added_tag_one.setText("")
                self.label_added_tag_two.setText("")
                self.add_tag_to_question_button.setText("Add tag >")
        else:
            wb = openpyxl.load_workbook(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
            sheet = wb.active
            current_question = int(self.label_question.text().replace(":", "")[19:])
            sheet["A" + str(current_question + 1)] = self.question_line_edit.text().capitalize()  # QUESTION
            sheet["D" + str(current_question + 1)] = self.answer_text_edit.toPlainText().capitalize()  # ANSWER

            if self.easy_check_box.isChecked():  # DIFFICULTY
                difficulty = 1
                sheet["H" + str(current_question + 1)] = difficulty
            elif self.medium_check_box.isChecked():
                difficulty = 3
                sheet["H" + str(current_question + 1)] = difficulty
            elif self.hard_check_box.isChecked():
                difficulty = 5
                sheet["H" + str(current_question + 1)] = difficulty
            else:
                difficulty = 7
                sheet["H" + str(current_question + 1)] = difficulty

            wb.save(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
            wb.close()
            self.generic_information("Successfully added",
                                     f"Question {current_question} was added correctly.\n\n"
                                     f"Question: {self.question_line_edit.text().capitalize()}\n\n"
                                     f"Answer: {self.answer_text_edit.toPlainText().capitalize()}\n\n"
                                     f"Tags: {self.label_added_tag_one.text()} / {self.label_added_tag_two.text()}"
                                     f"\n\nDifficulty: {self.difficulty_dict.get(difficulty)}")
            self.question_line_edit.clear()
            self.answer_text_edit.clear()
            self.wrong_answer_text_edit_1.clear()
            self.wrong_answer_text_edit_2.clear()
            self.wrong_answer_text_edit_3.clear()
            self.label_added_tag_one.setText("")
            self.label_added_tag_two.setText("")
            self.add_tag_to_question_button.setText("Add tag >")

            # Great way to access current questionnaire, could be used in a lot more places
            items = []
            for i in range(self.main_list_widget.count()):
                items.append(self.main_list_widget.item(i).text())
            current_questionnaire = items.index(self.label_project_title_questionnaire.text())
            self.set_active_questionnaire(current_questionnaire)

            self.mod_add_question_button.hide()
            self.add_question_button.show()

    def delete_question(self):
        # FETCHING DATA
        # -- If there is a currently selected questionnaire, other that default
        if self.label_project_title_questionnaire.text() != "Questionnaire Title":
            wb = openpyxl.load_workbook(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
            sheet = wb.active
            current_question = int(self.label_question.text().replace(":", "")[9:])
            # -- Fetch questions data from excel file
            question_list = []
            for cell in sheet["A"]:  # Adds all question values to a list
                if cell.value != "QUESTIONS" and cell.value is not None:
                    question_list.append(cell.value)
            # -- Present user a dialog to choose question to delete
            choice, ok = QtWidgets.QInputDialog.getItem(self.window,
                                                        "Delete Question",
                                                        "Select a question from the list to delete",
                                                        question_list, 0, False)
            # MODIFYING DATA
            # -- If user chooses a question from dialog, delete that question and raise success message
            if choice and ok:
                tag = sheet["B" + str(question_list.index(choice) + 2)].value
                # -- If there is no tag value in deleted row, just delete the row
                if tag is None or tag == "":
                    sheet.delete_rows((question_list.index(choice) + 2), 1)
                # -- If there is a tag value in to-be-deleted row, first move tag value to another unused row
                else:
                    # -- For every cell in B column, check for an empty space, then relocate tag value
                    for cell in sheet["B"]:
                        if cell.value != "TAGS" and cell.value is None:
                            sheet[cell.coordinate] = tag
                            break
                    sheet.delete_rows((question_list.index(choice) + 2), 1)
                # -- Raise success message
                self.generic_information("Question Deleted", f"The following question: \n\n{choice}\n\n"
                                                             f"And all of its items were deleted.")
            # -- Save and close modified data
            wb.save(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
            wb.close()

            # RESET ACTIVE QUESTIONNAIRE
            self.set_active_questionnaire(self.get_active_questionnaire_index())

    def add_tags(self):
        tag, ok = QtWidgets.QInputDialog.getText(self.window,
                                                 "Add tag", "Enter tag name", QtWidgets.QLineEdit.Normal)
        if tag and ok:
            wb = openpyxl.load_workbook(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
            sheet = wb.active
            sheet["B" + str(sheet.max_row + 1)] = tag.upper()  # High sus on this, might break tags in future
            self.tag_combo_box.addItem(tag.upper())
            wb.save(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
            wb.close()

    def remove_tags(self):
        wb = openpyxl.load_workbook(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
        sheet = wb.active
        tag_list = []
        for cell in sheet["B"]:
            if cell.value != "TAGS" and cell.value is not None:
                tag_list.append(cell.value)
        tag, ok = QtWidgets.QInputDialog.getItem(self.window, "Delete a tag",
                                                 "Please select a tag from the dropdown list to delete it.",
                                                 tag_list, 0, False)
        if tag and ok:
            index = tag_list.index(tag)
            for cell in sheet["B"]:
                if cell.value == tag:
                    sheet[cell.coordinate] = ""
                    self.tag_combo_box.removeItem(index)
        wb.save(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
        wb.close()

    def add_tag_to_question(self):
        wb = openpyxl.load_workbook(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
        sheet = wb.active
        # UI CHECK - Works using labels as reference
        current_tag_one = self.label_added_tag_one.text()
        current_tag_two = self.label_added_tag_two.text()
        tag_list = [self.tag_combo_box.itemText(i) for i in range(self.tag_combo_box.count())]
        # -- If user is modifying question, consider alternate question label, otherwise consider default
        # Very delicate, if program breaks, high sus
        if self.label_question.text().startswith("Mod"):
            current_question = int(self.label_question.text().replace(":", "")[19:])
        else:
            current_question = int(self.label_question.text().replace(":", "")[9:])
        # -- If tags from other questionnaire are present, erase them to allow new ones
        if current_tag_one != "" and current_tag_two != "":  # Reverse functionality, to remove tags
            self.label_added_tag_one.setText("")
            self.label_added_tag_two.setText("")
            self.add_tag_to_question_button.setText("Add tag >")
            sheet["C" + str(current_question + 1)].value = None
            wb.save(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
            wb.close()
            return None
        # -- If no tags have been added, add first one NOTE: First tag will always be filled first
        if self.label_added_tag_one.text() == "":
            self.label_added_tag_one.setText(self.tag_combo_box.currentText().upper())
            self.label_added_tag_one.show()
            sheet["C" + str(current_question + 1)] = self.tag_combo_box.currentText().upper()
            # -- If tag index is the last in the list, set current item to first, otherwise, change to next
            if tag_list.index(self.tag_combo_box.currentText()) < (len(tag_list) - 1):
                self.tag_combo_box.setCurrentIndex(tag_list.index(self.tag_combo_box.currentText()) + 1)
            else:
                self.tag_combo_box.setCurrentIndex(0)
        # -- If selected tag is already added, do not allow adding duplicate tag and raise message to user
        elif self.tag_combo_box.currentText().upper() == current_tag_one \
                or self.tag_combo_box.currentText().upper() == current_tag_two:
            self.generic_error("Repeated Tag", "You can't add the same tag twice to the same question.")
        # If a first tag has been added, but not a second, add the second one
        elif self.label_added_tag_one.text() != "" and self.label_added_tag_two.text() == "":
            self.label_added_tag_two.setText(self.tag_combo_box.currentText().upper())
            self.label_added_tag_two.show()
            # WRITING DATA TO EXCEL FILE
            if sheet["C" + str(current_question + 1)].value is not None:
                db_tag_one = sheet["C" + str(current_question + 1)].value
                sheet["C" + str(current_question + 1)] = db_tag_one + "//" + self.tag_combo_box.currentText().upper()
            self.add_tag_to_question_button.setText("Remove tags >")
        wb.save(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
        wb.close()

    def preview_questionnaire(self):
        if self.label_project_title_questionnaire.text() != "Questionnaire Title":
            # INSTANTIATION
            custom_table = CustomTableWidget()
            # FETCHING DATA
            wb = openpyxl.load_workbook(f"Questionnaire_{self.label_project_title_questionnaire.text()}.xlsx")
            sheet = wb.active
            custom_table.preview_table.setRowCount(int(self.label_number_of_questions.text()[16:]))
            # -- Fetch Questions, then add them to table
            row = 0
            for cell in sheet["A"]:
                if cell.value != "QUESTIONS" and cell.value is not None:
                    custom_table.preview_table.setItem(row, 0, QtWidgets.QTableWidgetItem(cell.value))
                    row += 1
                    if len(cell.value) > 30:
                        custom_table.preview_table.setRowHeight(row - 1, 60)
            # -- Fetch Tags, then add them to table
            row = 0
            for cell in sheet["C"]:
                if cell.value != "QUESTION TAGS":
                    custom_table.preview_table.setItem(row, 1, QtWidgets.QTableWidgetItem(cell.value))
                    row += 1
            # -- Fetch Answers, then add them to table
            row = 0
            for cell in sheet["D"]:
                if cell.value != "ANSWER" and cell.value is not None:
                    custom_table.preview_table.setItem(row, 2, QtWidgets.QTableWidgetItem(cell.value))
                    row += 1
                    if len(cell.value) > 30:
                        custom_table.preview_table.setRowHeight(row - 1, 60)
            # -- Fetch Optional Multiple Answer mode, then add them to table
            row = 0
            for cell in sheet["E"]:
                if cell.value != "WRONG ANSWER 1" and cell.value is not None:
                    custom_table.preview_table.setItem(row, 3, QtWidgets.QTableWidgetItem("✅"))
                    row += 1
                elif cell.value is None:
                    custom_table.preview_table.setItem(row, 3, QtWidgets.QTableWidgetItem("❌"))
                    row += 1
            wb.close()
            custom_table.__init__()  # Why does this work, needs research, might be unstable

    @staticmethod
    def generic_information(title="Information", text="Details", icon=QMessageBox.Information):
        parent = QtWidgets.QMessageBox()
        parent.setGeometry(800, 300, 500, 500)
        parent.setWindowTitle(title)
        parent.setText(text)
        parent.setIcon(icon)
        parent.exec_()

    @staticmethod
    def generic_error(title="Error", text="Error Details", icon=QMessageBox.Warning):
        parent = QtWidgets.QMessageBox()
        parent.setGeometry(800, 300, 500, 500)
        parent.setWindowTitle(title)
        parent.setText(text)
        parent.setIcon(icon)
        parent.exec_()

    @staticmethod
    def switch_to_quiz():
        questionnaireMainWindow.close()
        QuizMainWindow.show()


class CustomTableWidget(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        CustomTableWidget.setWindowTitle(self, "Preview Table")
        self.setGeometry(800, 400, 700, 300)
        self.gridLayout = QtWidgets.QGridLayout(self)
        self.frame = QtWidgets.QFrame(self)
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.frame)
        self.verticalLayout.setObjectName("verticalLayout")
        self.preview_table = QtWidgets.QTableWidget()
        self.preview_table.setMinimumSize(680, 280)
        self.preview_table.setAlternatingRowColors(True)
        item = QtWidgets.QTableWidgetItem()
        item2 = QtWidgets.QTableWidgetItem()
        item3 = QtWidgets.QTableWidgetItem()
        item4 = QtWidgets.QTableWidgetItem()
        self.preview_table.setColumnCount(4)
        self.preview_table.setRowCount(5)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.preview_table.setHorizontalHeaderItem(0, item)
        self.preview_table.setHorizontalHeaderItem(1, item2)
        self.preview_table.setHorizontalHeaderItem(2, item3)
        self.preview_table.setHorizontalHeaderItem(3, item4)

        item = self.preview_table.horizontalHeaderItem(0)
        item.setText("Question")
        item = self.preview_table.horizontalHeaderItem(1)
        item.setText("Tags")
        item = self.preview_table.horizontalHeaderItem(2)
        item.setText("Answer")
        item = self.preview_table.horizontalHeaderItem(3)
        item.setText("Wrong answers")

        self.preview_table.setColumnWidth(0, 220)
        self.preview_table.setColumnWidth(1, 100)
        self.preview_table.setColumnWidth(2, 220)
        self.preview_table.setColumnWidth(3, 70)
        self.preview_table.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        self.verticalLayout.addWidget(self.preview_table)
        self.gridLayout.addWidget(self.frame, 0, 0, 1, 1)
        self.show()


class QuizzerMainWindow(object):
    def __init__(self):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(483, 492)
        MainWindow.setGeometry(800, 300, 500, 500)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.home_frame = QtWidgets.QFrame(self.centralwidget)
        self.home_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.home_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.home_frame.setObjectName("home_frame")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.home_frame)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.pix_label_questionnaire = clickable_label.QClickableLabel(self.home_frame)
        self.pix_label_questionnaire.setMinimumSize(QtCore.QSize(215, 215))
        self.pix_label_questionnaire.setAutoFillBackground(False)
        self.pix_label_questionnaire.setFrameShape(QtWidgets.QFrame.Panel)
        self.pix_label_questionnaire.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.pix_label_questionnaire.setLineWidth(5)
        self.pix_label_questionnaire.setMidLineWidth(0)
        self.pix_label_questionnaire.setText("")
        self.pix_label_questionnaire.setPixmap(QtGui.QPixmap("questionnaire.gif"))
        self.pix_label_questionnaire.setAlignment(QtCore.Qt.AlignCenter)
        self.pix_label_questionnaire.setObjectName("pix_label_questionnaire")
        self.gridLayout_2.addWidget(self.pix_label_questionnaire, 1, 0, 1, 1)
        self.label_questionnaire = QtWidgets.QLabel(self.home_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_questionnaire.setFont(font)
        self.label_questionnaire.setAlignment(QtCore.Qt.AlignCenter)
        self.label_questionnaire.setObjectName("label_questionnaire")
        self.gridLayout_2.addWidget(self.label_questionnaire, 2, 0, 1, 1)
        self.label_quiz = QtWidgets.QLabel(self.home_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_quiz.setFont(font)
        self.label_quiz.setAlignment(QtCore.Qt.AlignCenter)
        self.label_quiz.setObjectName("label_quiz")
        self.gridLayout_2.addWidget(self.label_quiz, 2, 2, 1, 1)
        self.pix_label_quiz = clickable_label.QClickableLabel(self.home_frame)
        self.pix_label_quiz.setMinimumSize(QtCore.QSize(215, 215))
        self.pix_label_quiz.setFrameShape(QtWidgets.QFrame.Panel)
        self.pix_label_quiz.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.pix_label_quiz.setLineWidth(5)
        self.pix_label_quiz.setText("")
        self.pix_label_quiz.setPixmap(QtGui.QPixmap("quiz-icon.gif"))
        self.pix_label_quiz.setAlignment(QtCore.Qt.AlignCenter)
        self.pix_label_quiz.setObjectName("pix_label_quiz")
        self.gridLayout_2.addWidget(self.pix_label_quiz, 1, 2, 1, 1)
        self.label_home_title = QtWidgets.QLabel(self.home_frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Minimum)
        sizePolicy.setHorizontalStretch(100)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_home_title.sizePolicy().hasHeightForWidth())
        self.label_home_title.setSizePolicy(sizePolicy)
        self.label_home_title.setMinimumSize(QtCore.QSize(0, 100))
        font = QtGui.QFont()
        font.setPointSize(21)
        font.setBold(True)
        font.setItalic(False)
        font.setUnderline(False)
        font.setWeight(75)
        self.label_home_title.setFont(font)
        self.label_home_title.setAlignment(QtCore.Qt.AlignCenter)
        self.label_home_title.setObjectName("label_home_title")
        self.gridLayout_2.addWidget(self.label_home_title, 0, 0, 1, 3)
        self.label_questionnaire_description = QtWidgets.QLabel(self.home_frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(40)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_questionnaire_description.sizePolicy().hasHeightForWidth())
        self.label_questionnaire_description.setSizePolicy(sizePolicy)
        self.label_questionnaire_description.setMinimumSize(QtCore.QSize(0, 20))
        self.label_questionnaire_description.setWordWrap(True)
        self.label_questionnaire_description.setObjectName("label_questionnaire_description")
        self.gridLayout_2.addWidget(self.label_questionnaire_description, 3, 0, 1, 1)
        self.line = QtWidgets.QFrame(self.home_frame)
        self.line.setFrameShape(QtWidgets.QFrame.VLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.gridLayout_2.addWidget(self.line, 1, 1, 3, 1)
        self.label_quizzes_description = QtWidgets.QLabel(self.home_frame)
        self.label_quizzes_description.setObjectName("label_quizzes_description")
        self.gridLayout_2.addWidget(self.label_quizzes_description, 3, 2, 1, 1)
        self.gridLayout.addWidget(self.home_frame, 0, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 483, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.textify(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # Connections
        self.pix_label_questionnaire.clicked.connect(self.show_questionnaire)
        self.pix_label_quiz.clicked.connect(self.show_quiz)
        self.pix_label_questionnaire.clicked.connect(self.show_questionnaire)

    def textify(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Quizzer App"))
        self.label_questionnaire.setText(_translate("MainWindow", "Questionnaires"))
        self.label_quiz.setText(_translate("MainWindow", "Quizzes"))
        self.label_home_title.setText(_translate("MainWindow", "Quizzes App"))
        self.label_questionnaire_description.setText(_translate("MainWindow", "Create your own questionnaires and "
                                                                              "prepare them to transform into "
                                                                              "quizzes!"))
        self.label_quizzes_description.setText(_translate("MainWindow", "Test your own quizzes!"))

    def show_quiz(self):
        MainWindow.close()
        QuizMainWindow.show()

    def show_questionnaire(self):
        MainWindow.close()
        questionnaireMainWindow.show()


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    quizzer_ui = QuizzerMainWindow()
    quizzer_ui.__init__()
    MainWindow.show()
    QuizMainWindow = QtWidgets.QMainWindow()
    quiz_ui = UiQuizMainWindow()
    quiz_ui.__init__()
    questionnaireMainWindow = QtWidgets.QMainWindow()
    questionnaire_ui = UiQuestionnaireMainWindow()
    questionnaire_ui.__init__()
    custom_table = CustomTableWidget()
    custom_table.hide()

    sys.exit(app.exec_())
